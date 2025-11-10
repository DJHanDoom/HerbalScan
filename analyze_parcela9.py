import openpyxl

# Dados analisados das 10 subparcelas da Parcela 9
# Baseado nas imagens TimePhoto ordenadas por timestamp

parcela9_data = [
    # Subparcela 1 (TimePhoto_20251107_133034.jpg - índice 4107)
    (9, 1, 1, 'Gramínea Alta Verde', 30, 35, 'Erva'),
    (9, 1, 2, 'Gramínea Seca Palha', 45, 25, 'Erva'),
    (9, 1, 3, 'Gramínea Fina Verde', 15, 20, 'Erva'),

    # Subparcela 2 (TimePhoto_20251107_133249.jpg - índice 4108)
    (9, 2, 1, 'Gramínea Seca Palha', 60, 30, 'Erva'),
    (9, 2, 2, 'Gramínea Alta Verde', 20, 40, 'Erva'),
    (9, 2, 3, 'Gramínea Fina Verde', 10, 18, 'Erva'),

    # Subparcela 3 (TimePhoto_20251107_133446.jpg - índice 4140)
    (9, 3, 1, 'Gramínea Seca Palha', 55, 28, 'Erva'),
    (9, 3, 2, 'Gramínea Alta Verde', 25, 35, 'Erva'),
    (9, 3, 3, 'Gramínea Fina Verde', 10, 15, 'Erva'),

    # Subparcela 4 (TimePhoto_20251107_133807.jpg - índice 4119)
    (9, 4, 1, 'Gramínea Seca Palha', 40, 22, 'Erva'),
    (9, 4, 2, 'Solo Exposto', 35, 0, '-'),
    (9, 4, 3, 'Gramínea Baixa Roseta', 15, 8, 'Erva'),
    (9, 4, 4, 'Gramínea Fina Verde', 10, 12, 'Erva'),

    # Subparcela 5 (TimePhoto_20251107_134123.jpg - índice 4121)
    (9, 5, 1, 'Gramínea Seca Palha', 70, 25, 'Erva'),
    (9, 5, 2, 'Gramínea Alta Verde', 20, 30, 'Erva'),
    (9, 5, 3, 'Gramínea Fina Verde', 10, 15, 'Erva'),

    # Subparcela 6 (TimePhoto_20251107_134343.jpg - índice 4122)
    (9, 6, 1, 'Gramínea Seca Palha', 50, 30, 'Erva'),
    (9, 6, 2, 'Gramínea Alta Verde', 30, 38, 'Erva'),
    (9, 6, 3, 'Gramínea Fina Verde', 15, 20, 'Erva'),

    # Subparcela 7 (TimePhoto_20251107_134654.jpg - índice 4127)
    (9, 7, 1, 'Gramínea Seca Palha', 45, 28, 'Erva'),
    (9, 7, 2, 'Gramínea Alta Verde', 25, 35, 'Erva'),
    (9, 7, 3, 'Gramínea Fina Verde', 20, 18, 'Erva'),
    (9, 7, 4, 'Solo Exposto', 10, 0, '-'),

    # Subparcela 8 (TimePhoto_20251107_135003.jpg - índice 4126)
    (9, 8, 1, 'Gramínea Alta Verde', 40, 42, 'Erva'),
    (9, 8, 2, 'Gramínea Seca Palha', 35, 25, 'Erva'),
    (9, 8, 3, 'Gramínea Fina Verde', 15, 18, 'Erva'),
    (9, 8, 4, 'Erva Folha Larga', 10, 12, 'Erva'),

    # Subparcela 9 (TimePhoto_20251107_135459.jpg - índice 4155)
    (9, 9, 1, 'Gramínea Alta Verde', 45, 45, 'Erva'),
    (9, 9, 2, 'Gramínea Seca Palha', 35, 30, 'Erva'),
    (9, 9, 3, 'Gramínea Fina Verde', 15, 20, 'Erva'),

    # Subparcela 10 (TimePhoto_20251107_135623.jpg - índice 4136)
    (9, 10, 1, 'Gramínea Alta Verde', 50, 48, 'Erva'),
    (9, 10, 2, 'Gramínea Seca Palha', 30, 28, 'Erva'),
    (9, 10, 3, 'Gramínea Fina Verde', 15, 22, 'Erva'),
]

# Carregar planilha existente
wb = openpyxl.load_workbook(r'C:\Users\diogo\Documents\TRABALHO\dossel\subparcelasHerbaceas.xlsx')
ws = wb.active

# Encontrar última linha
last_row = ws.max_row

# Adicionar dados da parcela 9
current_row = last_row + 1
for data in parcela9_data:
    cn, sub, ind, nome, area, altura, forma = data
    ws.cell(row=current_row, column=1, value=cn if ind == 1 else None)  # CN só na primeira espécie
    ws.cell(row=current_row, column=2, value=sub if ind == 1 else None)  # sub só na primeira espécie
    ws.cell(row=current_row, column=3, value=ind)
    ws.cell(row=current_row, column=4, value=nome)
    ws.cell(row=current_row, column=5, value=area)
    ws.cell(row=current_row, column=6, value=altura)
    ws.cell(row=current_row, column=7, value=forma)
    current_row += 1

# Salvar planilha
wb.save(r'C:\Users\diogo\Documents\TRABALHO\dossel\subparcelasHerbaceas.xlsx')
print(f'Dados da Parcela 9 adicionados com sucesso!')
print(f'Total de registros adicionados: {len(parcela9_data)}')
print(f'Linha inicial: {last_row + 1}')
print(f'Linha final: {current_row - 1}')
