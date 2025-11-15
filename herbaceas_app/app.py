import os
import json
import zipfile
import io
import shutil
import sys
from flask import Flask, render_template, request, jsonify, send_file, Response, stream_with_context
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import base64
from pathlib import Path
import time
from prompt_templates import build_prompt, get_template_list, get_template_params, PROMPT_TEMPLATES

# Importações condicionais para diferentes IAs
try:
    import anthropic
    CLAUDE_AVAILABLE = True
except ImportError:
    CLAUDE_AVAILABLE = False

try:
    import openai
    GPT_AVAILABLE = True
    DEEPSEEK_AVAILABLE = True  # DeepSeek usa API compatível com OpenAI
except ImportError:
    GPT_AVAILABLE = False
    DEEPSEEK_AVAILABLE = False

try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

try:
    from openai import OpenAI
    QWEN_AVAILABLE = True  # Qwen (Alibaba) via API compatível
except ImportError:
    QWEN_AVAILABLE = False

try:
    import requests
    HUGGINGFACE_AVAILABLE = True
except ImportError:
    HUGGINGFACE_AVAILABLE = False

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg'}
app.config['DEFAULT_AI'] = os.environ.get('DEFAULT_AI', 'gemini')  # gemini como padrão

# Criar diretórios necessários
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('exports', exist_ok=True)

# Armazenamento em memória para dados da análise
analysis_data = {
    'parcelas': {},
    'especies_unificadas': {}
}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def decode_api_key(encoded_key):
    """
    Decodifica uma chave API que foi codificada em Base64 no frontend.
    Retorna a chave original ou None se inválida.
    """
    if not encoded_key:
        return None
    try:
        # Tentar decodificar Base64
        decoded = base64.b64decode(encoded_key).decode('utf-8')
        return decoded
    except Exception:
        # Se falhar, assumir que já é texto plano (compatibilidade)
        return encoded_key

def clean_json_response(text):
    """
    Limpa e repara texto JSON potencialmente malformado.
    Remove caracteres problemáticos e tenta extrair JSON válido.
    """
    import re
    
    # Remover espaços em branco extras
    text = text.strip()
    
    # Remover blocos de código markdown
    if text.startswith('```json'):
        text = text[7:]
    elif text.startswith('```'):
        text = text[3:]
    
    if text.endswith('```'):
        text = text[:-3]
    
    text = text.strip()
    
    # Tentar encontrar JSON entre chaves
    start = text.find('{')
    end = text.rfind('}')
    
    if start != -1 and end != -1 and end > start:
        text = text[start:end+1]
    
    # Substituir aspas simples por duplas em chaves (comum em respostas de IA)
    # Padrão para encontrar chaves entre aspas simples
    text = re.sub(r"'([a-zA-Z_][a-zA-Z0-9_]*)'(\s*):", r'"\1"\2:', text)
    
    # 🔧 FIX: Corrigir vírgulas ausentes após valores em objetos
    # Padrão: "key": valor\n  "nextkey" → "key": valor,\n  "nextkey"
    text = re.sub(r'(":\s*(?:"[^"]*"|[0-9.]+|true|false|null))\s*\n\s*"', r'\1,\n  "', text)
    
    # 🔧 FIX: Remover vírgulas antes de fechar array/objeto
    text = re.sub(r',(\s*[}\]])', r'\1', text)
    
    # 🔧 FIX: Adicionar vírgulas entre objetos em array
    # Padrão: }\n  { → },\n  {
    text = re.sub(r'}\s*\n\s*{', r'},\n    {', text)
    
    return text.strip()

def fix_malformed_json(text, json_error):
    """
    Tenta corrigir JSON malformado baseado no erro específico.
    """
    import re
    
    error_msg = str(json_error)
    
    # Caso 1: String não terminada (JSON truncado)
    if "Unterminated string" in error_msg:
        print("🔧 Detectado: String não terminada (JSON possivelmente truncado)")
        # Tentar fechar a string e o objeto
        text = text.rstrip()
        # Se não termina com aspas, adicionar
        if not text.endswith('"'):
            text += '"'
        # Se não termina com }, adicionar estrutura de fechamento
        if not text.endswith('}') and not text.endswith(']'):
            # Contar { e } para balancear
            open_braces = text.count('{')
            close_braces = text.count('}')
            open_brackets = text.count('[')
            close_brackets = text.count(']')
            
            # Adicionar vírgulas e fechamentos necessários
            if ',' not in text[-20:]:  # Se não tem vírgula recente
                text += ','
            text += '\n    "cobertura": 0,'
            text += '\n    "altura": 0,'
            text += '\n    "forma_vida": "Erva"'
            text += '\n  }' * (open_braces - close_braces)
            text += '\n]' * (open_brackets - close_brackets)
        
        print(f"✓ String fechada, estrutura balanceada")
        return text
    
    # Caso 2: Falta vírgula entre campos
    if "line" in error_msg and "column" in error_msg:
        # Converter para lista de linhas
        lines = text.split('\n')
        
        # Encontrar posição do erro
        match = re.search(r'line (\d+) column (\d+)', error_msg)
        if match:
            line_num = int(match.group(1)) - 1  # 0-indexed
            col_num = int(match.group(2)) - 1
            
            print(f"🔍 Erro na linha {line_num + 1}, coluna {col_num + 1}")
            
            if line_num < len(lines):
                error_line = lines[line_num]
                print(f"🔍 Linha problemática: {error_line}")
                
                # Caso 2a: Falta vírgula após valor
                if "Expecting ',' delimiter" in error_msg:
                    # Verificar se próxima linha começa com aspas (novo campo)
                    if line_num + 1 < len(lines):
                        next_line = lines[line_num + 1].strip()
                        if next_line.startswith('"'):
                            # Adicionar vírgula no final da linha atual
                            lines[line_num] = error_line.rstrip() + ','
                            print(f"✓ Adicionada vírgula após: {error_line[:50]}")
                
                # Caso 2b: Vírgula extra antes de } ou ]
                elif "Expecting '}'" in error_msg or "Expecting ']'" in error_msg:
                    if error_line.rstrip().endswith(','):
                        lines[line_num] = error_line.rstrip()[:-1]
                        print(f"✓ Removida vírgula extra")
        
        # Reconstruir texto
        text = '\n'.join(lines)
    
    return text

def validate_and_filter_results(analysis_result, template_config=None):
    """
    Valida e filtra os resultados da análise de acordo com as configurações do template.
    Remove categorias não solicitadas e ajusta número de morfotipos se necessário.
    
    Args:
        analysis_result: dict com 'especies' array
        template_config: dict com 'template' e 'params'
    
    Returns:
        dict: resultado filtrado e validado
    """
    if not template_config or 'especies' not in analysis_result:
        return analysis_result
    
    print(f"🔍 validate_and_filter_results - Entrada: {len(analysis_result.get('especies', []))} espécies")
    
    params = template_config.get('params', {})
    if not params:
        # Obter params do template default
        template_name = template_config.get('template', 'default')
        if template_name in PROMPT_TEMPLATES:
            params = PROMPT_TEMPLATES[template_name]['params'].copy()
        else:
            return analysis_result
    
    print(f"🔍 validate_and_filter_results - INÍCIO")
    print(f"   Total espécies RECEBIDAS: {len(analysis_result.get('especies', []))}")
    if analysis_result.get('especies'):
        print(f"   Lista espécies:")
        for i, esp in enumerate(analysis_result['especies'][:5], 1):
            print(f"      {i}. '{esp.get('apelido', 'N/A')}' (cob: {esp.get('cobertura', 0)}%)")
    print(f"🔍 Parâmetros: include_soil={params.get('include_soil')}, include_litter={params.get('include_litter')}")
    print(f"🔍 Limites: min={params.get('min_species', 1)}, max={params.get('max_species', 12)}")
    
    especies_filtradas = []
    
    for esp in analysis_result['especies']:
        apelido = esp.get('apelido', '').lower()
        
        # Filtrar solo exposto se não deve ser incluído
        if not params.get('include_soil', True):
            if any(termo in apelido for termo in ['solo exposto', 'solo nu', 'bare soil', 'exposed soil']):
                print(f"  ⊗ Filtrado (solo): {esp.get('apelido')}")
                continue
        
        # Filtrar serapilheira se não deve ser incluída
        if not params.get('include_litter', True):
            if any(termo in apelido for termo in ['serapilheira', 'folhiço', 'litter', 'material morto', 'detritos']):
                print(f"  ⊗ Filtrado (serapilheira): {esp.get('apelido')}")
                continue
        
        print(f"  ✓ Aceito: {esp.get('apelido')}")
        especies_filtradas.append(esp)
    
    print(f"🔍 Após filtragem: {len(especies_filtradas)} espécies")
    
    # Verificar limites de espécies
    max_especies = params.get('max_species', 12)
    min_especies = params.get('min_species', 1)  # Padrão: aceitar pelo menos 1 espécie
    
    print(f"🔍 Limites: min={min_especies}, max={max_especies}")
    
    if len(especies_filtradas) > max_especies:
        print(f"⚠️ Reduzindo de {len(especies_filtradas)} para {max_especies} morfotipos (limite configurado)")
        # Ordenar por cobertura e manter os mais relevantes
        especies_filtradas.sort(key=lambda x: x.get('cobertura', 0), reverse=True)
        especies_filtradas = especies_filtradas[:max_especies]
    
    # VALIDAÇÃO CRÍTICA: Rejeitar se não houver espécies suficientes
    if len(especies_filtradas) == 0:
        print(f"❌ VALIDAÇÃO FALHOU: 0 espécies após filtragem!")
        raise ValueError(
            f"❌ ERRO: Nenhuma espécie detectada após filtragem. "
            f"IA retornou {len(analysis_result.get('especies', []))} morfotipos originalmente. "
            f"Verifique se a imagem contém vegetação visível ou ajuste as configurações."
        )
    
    # VALIDAÇÃO: Rejeitar se houver menos espécies que o mínimo configurado
    if len(especies_filtradas) < min_especies:
        if len(analysis_result.get('especies', [])) < min_especies:
            # IA retornou menos espécies do que o mínimo - REJEITAR
            print(f"❌ VALIDAÇÃO FALHOU: IA retornou {len(analysis_result['especies'])} < mínimo {min_especies}")
            raise ValueError(
                f"❌ ERRO: IA detectou apenas {len(analysis_result['especies'])} morfotipos "
                f"(mínimo configurado: {min_especies}). "
                f"Tentando próximo modelo ou clique em 'Reanalisar' para tentar novamente."
            )
        else:
            # Filtragem removeu espécies demais - REJEITAR
            raise ValueError(
                f"❌ ERRO: Apenas {len(especies_filtradas)} morfotipos restaram após filtragem "
                f"(mínimo: {min_especies}). "
                f"IA retornou {len(analysis_result['especies'])} originalmente. "
                f"Configure para incluir solo/serapilheira ou ajuste o mínimo de espécies."
            )
    
    # Calcular cobertura total
    total_cobertura = sum(esp.get('cobertura', 0) for esp in especies_filtradas)
    
    # 🔧 OPCIONAL: Normalizar coberturas para somar 100%
    normalize_coverage = params.get('normalize_coverage', False)
    
    if normalize_coverage and total_cobertura > 0 and abs(total_cobertura - 100) > 5:
        # Recalcular proporcionalmente para somar 100%
        fator = 100 / total_cobertura
        for esp in especies_filtradas:
            esp['cobertura'] = round(esp['cobertura'] * fator, 1)
        print(f"✓ Coberturas normalizadas: {total_cobertura}% → 100%")
        total_cobertura = 100
    else:
        # Manter valores reais (padrão)
        print(f"✓ Coberturas mantidas reais: total {total_cobertura}%")
    
    print(f"✓ Validação concluída: {len(especies_filtradas)} morfotipos (cobertura total: {total_cobertura}%)")
    
    analysis_result['especies'] = especies_filtradas
    return analysis_result

def get_analysis_prompt(template_name="default", custom_params=None, custom_prompt=None):
    """
    Retorna o prompt para análise de imagens baseado em template ou prompt customizado
    
    Args:
        template_name: nome do template (default, regeneracao, carbono, etc)
        custom_params: dict com parâmetros customizados para sobrescrever
        custom_prompt: prompt editado manualmente pelo usuário (sobrescreve template)
    
    Returns:
        str: prompt formatado
    """
    # Se houver prompt customizado (editado manualmente), usar ele diretamente
    if custom_prompt:
        print("📝 Usando prompt editado manualmente pelo usuário")
        return custom_prompt
    
    return build_prompt(template_name, custom_params)

def analyze_image_with_claude(image_path, api_key=None, model_version=None, template_config=None):
    """Analisa uma imagem usando Claude API"""
    try:
        # Usar API key fornecida ou variável de ambiente
        key = api_key or os.environ.get("ANTHROPIC_API_KEY")
        if not key:
            raise Exception("API key do Claude não configurada")

        client = anthropic.Anthropic(api_key=key)

        with open(image_path, "rb") as image_file:
            image_data = base64.b64encode(image_file.read()).decode("utf-8")

        ext = image_path.split('.')[-1].lower()
        media_type = f"image/{ext}" if ext != 'jpg' else "image/jpeg"

        # Gerar prompt baseado em template ou prompt customizado
        if template_config:
            prompt = get_analysis_prompt(
                template_config.get('template', 'default'),
                template_config.get('params'),
                template_config.get('customPrompt')  # Prompt editado manualmente
            )
        else:
            prompt = get_analysis_prompt()

        # Usar versão específica se fornecida, caso contrário tentar várias
        if model_version:
            model_names = [model_version]
            print(f"Usando versão específica do Claude: {model_version}")
        else:
            # Tentar diferentes modelos Claude em ordem de preferência
            # Modelos válidos Claude (2025 - verificado no console Anthropic)
            model_names = [
                "claude-opus-4-1-20250805",    # Opus 4.1 (mais poderoso!)
                "claude-sonnet-4-5-20250929",  # Sonnet 4.5 (latest)
                "claude-opus-4-20250514",      # Opus 4
                "claude-sonnet-4-20250514",    # Sonnet 4
                "claude-haiku-4-5-20251001",   # Haiku 4.5 (fastest)
                "claude-3-5-haiku-20241022",   # Haiku 3.5
                "claude-3-haiku-20240307"      # Haiku 3 (fallback)
            ]

        last_error = None
        for model_name in model_names:
            try:
                print(f"Tentando modelo Claude: {model_name}")
                
                message = client.messages.create(
                    model=model_name,
                    max_tokens=2000,
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {
                                    "type": "image",
                                    "source": {
                                        "type": "base64",
                                        "media_type": media_type,
                                        "data": image_data,
                                    },
                                },
                                {
                                    "type": "text",
                                    "text": prompt
                                }
                            ],
                        }
                    ],
                )

                response_text = message.content[0].text.strip()
                
                print(f"📝 Claude retornou {len(response_text)} caracteres")
                
                # Limpar resposta usando helper function
                response_text = clean_json_response(response_text)
                
                try:
                    result = json.loads(response_text)
                    print(f"✓ JSON parseado com sucesso")
                    print(f"📊 Espécies detectadas pelo Claude: {len(result.get('especies', []))}")
                    if result.get('especies'):
                        print(f"   Primeira espécie: {result['especies'][0].get('apelido', 'N/A')}")
                except json.JSONDecodeError as json_err:
                    print(f"⚠️ Erro ao parsear JSON do Claude")
                    print(f"Posição do erro: {json_err}")
                    print(f"Texto recebido (primeiros 500 chars):")
                    print(response_text[:500])
                    raise Exception(f"Resposta JSON inválida: {str(json_err)}")
                
                print(f"🔍 Iniciando validação com min_species={template_config.get('params', {}).get('min_species', 1)}")
                
                # Validar e filtrar resultados de acordo com configuração
                result = validate_and_filter_results(result, template_config)
                
                print(f"✓ Sucesso com modelo: {model_name}")
                return result

            except Exception as e:
                last_error = e
                error_str = str(e)
                print(f"✗ Falha com {model_name}: {error_str[:200]}")
                
                # Verificar se é erro de validação (poucas espécies) ou outro erro recuperável
                is_validation_error = "morfotipos" in error_str or "espécies" in error_str.lower()
                is_safety_block = "bloqueado" in error_str.lower() or "segurança" in error_str.lower()
                
                # Se versão específica foi solicitada
                if model_version:
                    # EXCEÇÃO: Se for erro de validação ou bloqueio, tentar outros modelos
                    if is_validation_error or is_safety_block:
                        print(f"⚠️ Erro recuperável detectado - tentando outros modelos Claude automaticamente")
                        # Adicionar todos os outros modelos à lista para tentar
                        if len(model_names) == 1:  # Se só tinha 1 modelo (o específico)
                            print("⚠️ Expandindo busca para todos os modelos Claude disponíveis")
                            model_names.extend([
                                "claude-opus-4-1-20250805",
                                "claude-sonnet-4-5-20250929",
                                "claude-opus-4-20250514",
                                "claude-sonnet-4-20250514",
                                "claude-haiku-4-5-20251001",
                                "claude-3-5-haiku-20241022",
                                "claude-3-haiku-20240307"
                            ])
                        continue  # Tentar próximo modelo
                    else:
                        # Para outros erros, falhar imediatamente
                        raise Exception(f"Falha com versão específica {model_version}: {error_str}")
                continue

        # Se nenhum modelo funcionou, fornecer mensagem clara
        error_msg = f"Nenhum modelo Claude conseguiu detectar espécies suficientes. Último erro: {str(last_error)}"
        print(f"❌ {error_msg}")
        raise Exception(error_msg)

    except Exception as e:
        error_msg = str(e)
        print(f"\n❌ ERRO NA ANÁLISE COM CLAUDE:")
        print(f"Tipo: {type(e).__name__}")
        print(f"Mensagem: {error_msg}")
        import traceback
        traceback.print_exc()
        print("=" * 50)

        return {
            "especies": [
                {
                    "apelido": "Erro na análise",
                    "cobertura": 0,
                    "altura": 0,
                    "forma_vida": "-",
                    "erro": error_msg
                }
            ]
        }

def analyze_image_with_gpt4(image_path, api_key=None, template_config=None):
    """Analisa uma imagem usando GPT-4 Vision"""
    try:
        key = api_key or os.environ.get("OPENAI_API_KEY")
        if not key:
            raise Exception("API key do GPT-4 não configurada")

        client = openai.OpenAI(api_key=key)
        
        # Configuração de template padrão
        if template_config is None:
            template_config = {'template': 'default', 'params': None}
        
        # Gerar prompt usando template ou prompt customizado (GPT-4)
        template_name = template_config.get('template', 'default')
        custom_params = template_config.get('params')
        custom_prompt = template_config.get('customPrompt')
        prompt_text = get_analysis_prompt(template_name, custom_params, custom_prompt)
        print(f"Usando template: {template_name}")

        with open(image_path, "rb") as image_file:
            image_data = base64.b64encode(image_file.read()).decode("utf-8")

        ext = image_path.split('.')[-1].lower()
        media_type = f"image/{ext}" if ext != 'jpg' else "image/jpeg"

        # Tentar múltiplos modelos GPT-4 com suporte a visão
        models_to_try = ["gpt-4o", "gpt-4-turbo", "gpt-4-vision-preview"]
        last_error = None

        for model_name in models_to_try:
            try:
                print(f"Tentando modelo: {model_name}")
                response = client.chat.completions.create(
                    model=model_name,
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {
                                    "type": "image_url",
                                    "image_url": {
                                        "url": f"data:{media_type};base64,{image_data}"
                                    }
                                },
                                {
                                    "type": "text",
                                    "text": prompt_text
                                }
                            ]
                        }
                    ],
                    max_tokens=2000
                )
                print(f"✓ Sucesso com modelo: {model_name}")
                break  # Sucesso, sair do loop
            except Exception as model_error:
                print(f"✗ Falha com {model_name}: {str(model_error)}")
                last_error = model_error
                continue
        else:
            # Se nenhum modelo funcionou
            raise Exception(f"Nenhum modelo GPT-4 disponível. Último erro: {str(last_error)}")

        response_text = response.choices[0].message.content.strip()
        if response_text.startswith('```'):
            lines = response_text.split('\n')
            response_text = '\n'.join(lines[1:-1])

        result = json.loads(response_text)
        
        # Validar e filtrar resultados de acordo com configuração
        result = validate_and_filter_results(result, template_config)
        
        return result

    except Exception as e:
        error_msg = str(e)
        print(f"\n❌ ERRO NA ANÁLISE COM GPT-4:")
        print(f"Tipo: {type(e).__name__}")
        print(f"Mensagem: {error_msg}")
        import traceback
        print(f"Traceback completo:")
        traceback.print_exc()
        print("=" * 50)

        return {
            "especies": [
                {
                    "apelido": "Erro na análise",
                    "cobertura": 0,
                    "altura": 0,
                    "forma_vida": "-",
                    "erro": error_msg
                }
            ]
        }

def analyze_image_with_gemini(image_path, api_key=None, model_version=None, template_config=None):
    """Analisa uma imagem usando Gemini"""
    try:
        key = api_key or os.environ.get("GOOGLE_API_KEY")
        if not key:
            raise Exception("API key do Gemini não configurada")

        genai.configure(api_key=key)
        
        # Configuração de template padrão
        if template_config is None:
            template_config = {'template': 'default', 'params': None}
        
        # Gerar prompt usando template ou prompt customizado (Gemini)
        template_name = template_config.get('template', 'default')
        custom_params = template_config.get('params')
        custom_prompt = template_config.get('customPrompt')
        prompt_text = get_analysis_prompt(template_name, custom_params, custom_prompt)
        print(f"Usando template: {template_name}")

        # Se um modelo específico foi fornecido, use-o primeiro
        # Caso contrário, use a lista de fallback
        if model_version:
            model_names = [model_version]
            print(f"Usando modelo Gemini específico: {model_version}")
        else:
            # Modelos Gemini 2.x (modelos 1.5 foram descontinuados)
            # Usando aliases automáticos que sempre apontam para versões mais recentes
            model_names = [
                'gemini-flash-latest',      # Alias para gemini-2.5-flash
                'gemini-2.5-flash',          # Versão rápida e eficiente
                'gemini-2.0-flash',          # Versão anterior estável
                'gemini-pro-latest'          # Alias para gemini-2.5-pro
            ]

        model = None
        last_error = None

        for model_name in model_names:
            try:
                print(f"Tentando modelo Gemini: {model_name}")
                
                # Configurar modelo com geração de conteúdo
                generation_config = {
                    "temperature": 0.4,
                    "top_p": 0.95,
                    "top_k": 40,
                    "max_output_tokens": 8192,  # 🔧 FIX: Aumentado para evitar JSON truncado
                }
                
                # Configurações de segurança mais permissivas para análise de vegetação
                safety_settings = [
                    {
                        "category": "HARM_CATEGORY_HARASSMENT",
                        "threshold": "BLOCK_NONE"
                    },
                    {
                        "category": "HARM_CATEGORY_HATE_SPEECH",
                        "threshold": "BLOCK_NONE"
                    },
                    {
                        "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
                        "threshold": "BLOCK_NONE"
                    },
                    {
                        "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
                        "threshold": "BLOCK_NONE"
                    }
                ]
                
                model = genai.GenerativeModel(
                    model_name=model_name,
                    generation_config=generation_config,
                    safety_settings=safety_settings
                )

                # Abrir imagem com PIL
                from PIL import Image
                img = Image.open(image_path)

                # Gerar resposta
                response = model.generate_content([
                    prompt_text,
                    img
                ])

                # Verificar se houve bloqueio por segurança
                if not response.candidates or not response.candidates[0].content.parts:
                    finish_reason = response.candidates[0].finish_reason if response.candidates else None
                    if finish_reason == 2:  # SAFETY
                        raise Exception("Conteúdo bloqueado por filtros de segurança. Tente outra imagem ou modelo.")
                    elif finish_reason == 3:  # RECITATION
                        raise Exception("Conteúdo bloqueado por violação de direitos autorais.")
                    else:
                        raise Exception(f"Resposta vazia do modelo (finish_reason: {finish_reason})")

                # Processar resposta
                response_text = response.text.strip()
                
                print(f"📝 Gemini {model_name} retornou {len(response_text)} caracteres")
                print(f"📝 Primeiros 200 chars: {response_text[:200]}")
                
                # Limpar resposta usando helper function
                response_text = clean_json_response(response_text)
                
                print(f"📝 Após limpeza: {len(response_text)} caracteres")
                
                # Tentar parsear JSON
                try:
                    result = json.loads(response_text)
                    print(f"✓ JSON parseado com sucesso")
                    print(f"✓ Número de espécies na resposta: {len(result.get('especies', []))}")
                    if result.get('especies'):
                        print(f"✓ Primeira espécie: {result['especies'][0].get('apelido', 'SEM NOME')}")
                except json.JSONDecodeError as json_err:
                    print(f"⚠️ Erro ao parsear JSON do Gemini")
                    print(f"Posição do erro: {json_err}")
                    print(f"Texto recebido (primeiros 500 chars):")
                    print(response_text[:500])
                    print(f"Texto recebido (últimos 200 chars):")
                    print(response_text[-200:])
                    
                    # 🔧 FIX: Tentar correção automática avançada
                    print("🔧 Tentando correção automática do JSON...")
                    try:
                        response_text = fix_malformed_json(response_text, json_err)
                        result = json.loads(response_text)
                        print(f"✓ JSON corrigido e parseado com sucesso!")
                    except Exception as fix_err:
                        print(f"❌ Correção automática falhou: {str(fix_err)}")
                        raise Exception(f"Resposta JSON inválida: {str(json_err)}")
                
                print(f"🔍 Antes da validação: {len(result.get('especies', []))} espécies")
                
                # Validar e filtrar resultados de acordo com configuração
                result = validate_and_filter_results(result, template_config)
                
                print(f"✓ Após validação: {len(result.get('especies', []))} espécies")
                print(f"✓ Sucesso com modelo: {model_name}")
                return result

            except Exception as e:
                last_error = e
                error_str = str(e)
                print(f"✗ Falha com {model_name}: {error_str[:200]}")
                
                # Verificar se é bloqueio de segurança (falso positivo comum em vegetação)
                is_safety_block = "segurança" in error_str or "bloqueado" in error_str or "finish_reason" in error_str
                
                # Se modelo específico foi solicitado e falhou
                if model_version:
                    # EXCEÇÃO: Se for bloqueio de segurança, tentar outros modelos automaticamente
                    if is_safety_block:
                        print("⚠️ Bloqueio de segurança detectado - tentando outros modelos Gemini automaticamente")
                        # Adicionar todos os outros modelos à lista para tentar
                        if len(model_names) == 1:  # Se só tinha 1 modelo (o específico)
                            print("⚠️ Expandindo busca para todos os modelos Gemini disponíveis")
                            model_names.extend([
                                "gemini-2.0-flash-exp",
                                "gemini-1.5-pro-latest",
                                "gemini-1.5-flash-latest",
                                "gemini-1.5-flash"
                            ])
                        continue  # Tentar próximo modelo
                    else:
                        # Para outros erros, falhar imediatamente
                        raise Exception(f"Falha com modelo específico {model_version}: {error_str}")
                
                # Se for erro de modelo não encontrado, tentar próximo
                if "404" in error_str or "not found" in error_str.lower():
                    continue
                # Se for erro de API key, não tentar outros modelos
                elif "api" in error_str.lower() and "key" in error_str.lower():
                    break
                # Se for bloqueio de segurança, tentar próximo modelo
                elif is_safety_block:
                    continue
                # Outros erros, tentar próximo
                continue

        # Se nenhum modelo funcionou, fornecer mensagem clara
        error_msg = f"Nenhum modelo Gemini conseguiu detectar espécies suficientes. Último erro: {str(last_error)}"
        print(f"❌ {error_msg}")
        raise Exception(error_msg)

    except Exception as e:
        error_msg = str(e)
        print(f"\n❌ ERRO NA ANÁLISE COM GEMINI:")
        print(f"Tipo: {type(e).__name__}")
        print(f"Mensagem: {error_msg}")
        import traceback
        print(f"Traceback completo:")
        traceback.print_exc()
        print("=" * 50)

        # Mensagem mais amigável
        user_msg = "Erro na análise com Gemini"
        if "404" in error_msg or "not found" in error_msg:
            user_msg = "Modelo Gemini não disponível"
        elif "API key" in error_msg or "invalid" in error_msg.lower():
            user_msg = "API key do Gemini inválida ou expirada"
        elif "quota" in error_msg.lower() or "limit" in error_msg.lower():
            user_msg = "Limite do Gemini atingido. Aguarde 1 minuto"

        return {
            "especies": [
                {
                    "apelido": user_msg,
                    "cobertura": 0,
                    "altura": 0,
                    "forma_vida": "-",
                    "erro": error_msg
                }
            ]
        }

def analyze_image_with_deepseek(image_path, api_key=None, template_config=None):
    """Analisa uma imagem usando DeepSeek (API compatível com OpenAI) - GRATUITO"""
    try:
        key = api_key or os.environ.get("DEEPSEEK_API_KEY")
        if not key:
            raise Exception("API key do DeepSeek não configurada")

        # Configuração de template padrão
        if template_config is None:
            template_config = {'template': 'default', 'params': None}
        
        # Gerar prompt usando template ou prompt customizado (DeepSeek)
        template_name = template_config.get('template', 'default')
        custom_params = template_config.get('params')
        custom_prompt = template_config.get('customPrompt')
        prompt_text = get_analysis_prompt(template_name, custom_params, custom_prompt)
        print(f"Usando template: {template_name}")

        # DeepSeek usa API compatível com OpenAI
        client = openai.OpenAI(
            api_key=key,
            base_url="https://api.deepseek.com/v1"
        )

        with open(image_path, "rb") as image_file:
            image_data = base64.standard_b64encode(image_file.read()).decode("utf-8")

        ext = image_path.split('.')[-1].lower()
        media_type = f"image/{ext}" if ext != 'jpg' else "image/jpeg"

        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{media_type};base64,{image_data}"
                            }
                        },
                        {
                            "type": "text",
                            "text": prompt_text
                        }
                    ]
                }
            ],
            max_tokens=2000
        )

        response_text = response.choices[0].message.content.strip()
        if response_text.startswith('```'):
            lines = response_text.split('\n')
            response_text = '\n'.join(lines[1:-1])

        result = json.loads(response_text)
        
        # Validar e filtrar resultados de acordo com configuração
        result = validate_and_filter_results(result, template_config)
        
        return result

    except Exception as e:
        print(f"Erro na análise com DeepSeek: {str(e)}")
        return {
            "especies": [{
                "apelido": "Erro DeepSeek",
                "cobertura": 0,
                "altura": 0,
                "forma_vida": "-",
                "erro": str(e)
            }]
        }

def analyze_image_with_qwen(image_path, api_key=None, template_config=None):
    """Analisa uma imagem usando Qwen (Alibaba) - GRATUITO"""
    try:
        key = api_key or os.environ.get("QWEN_API_KEY") or os.environ.get("DASHSCOPE_API_KEY")
        if not key:
            raise Exception("API key do Qwen não configurada")

        # Configuração de template padrão
        if template_config is None:
            template_config = {'template': 'default', 'params': None}
        
        # Gerar prompt usando template ou prompt customizado (Qwen)
        template_name = template_config.get('template', 'default')
        custom_params = template_config.get('params')
        custom_prompt = template_config.get('customPrompt')
        prompt_text = get_analysis_prompt(template_name, custom_params, custom_prompt)
        print(f"Usando template: {template_name}")

        # Qwen via DashScope (Alibaba Cloud)
        client = openai.OpenAI(
            api_key=key,
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"
        )

        with open(image_path, "rb") as image_file:
            image_data = base64.standard_b64encode(image_file.read()).decode("utf-8")

        ext = image_path.split('.')[-1].lower()
        media_type = f"image/{ext}" if ext != 'jpg' else "image/jpeg"

        response = client.chat.completions.create(
            model="qwen-vl-max",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{media_type};base64,{image_data}"
                            }
                        },
                        {
                            "type": "text",
                            "text": prompt_text
                        }
                    ]
                }
            ],
            max_tokens=2000
        )

        response_text = response.choices[0].message.content.strip()
        if response_text.startswith('```'):
            lines = response_text.split('\n')
            response_text = '\n'.join(lines[1:-1])

        result = json.loads(response_text)
        
        # Validar e filtrar resultados de acordo com configuração
        result = validate_and_filter_results(result, template_config)
        
        return result

    except Exception as e:
        print(f"Erro na análise com Qwen: {str(e)}")
        return {
            "especies": [{
                "apelido": "Erro Qwen",
                "cobertura": 0,
                "altura": 0,
                "forma_vida": "-",
                "erro": str(e)
            }]
        }

def analyze_image_with_huggingface(image_path, api_key=None, template_config=None):
    """Analisa uma imagem usando modelos da Hugging Face - GRATUITO"""
    try:
        key = api_key or os.environ.get("HUGGINGFACE_API_KEY")
        if not key:
            raise Exception("API key do Hugging Face não configurada")

        # Configuração de template padrão
        if template_config is None:
            template_config = {'template': 'default', 'params': None}
        
        # Gerar prompt usando template ou prompt customizado (HuggingFace)
        template_name = template_config.get('template', 'default')
        custom_params = template_config.get('params')
        custom_prompt = template_config.get('customPrompt')
        prompt_text = get_analysis_prompt(template_name, custom_params, custom_prompt)
        print(f"Usando template: {template_name}")

        # Usar modelo de visão da Hugging Face (ex: LLaVA)
        with open(image_path, "rb") as image_file:
            image_data = image_file.read()

        headers = {"Authorization": f"Bearer {key}"}

        # Tentar diferentes modelos
        models = [
            "llava-hf/llava-1.5-7b-hf",
            "Salesforce/blip-image-captioning-large",
            "nlpconnect/vit-gpt2-image-captioning"
        ]

        for model_id in models:
            try:
                api_url = f"https://api-inference.huggingface.co/models/{model_id}"

                payload = {
                    "inputs": prompt_text
                }

                response = requests.post(
                    api_url,
                    headers=headers,
                    files={"file": image_data},
                    data=payload,
                    timeout=30
                )

                if response.status_code == 200:
                    result = response.json()
                    # Processar resposta e converter para formato esperado
                    # Nota: pode precisar de parsing customizado dependendo do modelo
                    return {
                        "especies": [{
                            "apelido": "Análise HuggingFace",
                            "cobertura": 50,
                            "altura": 30,
                            "forma_vida": "Erva",
                            "nota": "Modelo gratuito - requer validação manual"
                        }]
                    }
            except Exception as e:
                print(f"Erro com modelo {model_id}: {str(e)}")
                continue

        raise Exception("Nenhum modelo HuggingFace respondeu")

    except Exception as e:
        print(f"Erro na análise com HuggingFace: {str(e)}")
        return {
            "especies": [{
                "apelido": "Erro HuggingFace",
                "cobertura": 0,
                "altura": 0,
                "forma_vida": "-",
                "erro": str(e)
            }]
        }

def analyze_image_with_ai(image_path, ai_model='claude', api_key=None, gemini_version=None, claude_version=None, template_config=None):
    """
    Analisa imagem com a IA selecionada
    
    Args:
        image_path: caminho da imagem
        ai_model: modelo de IA a usar
        api_key: chave API
        gemini_version: versão do Gemini (se aplicável)
        claude_version: versão do Claude (se aplicável)
        template_config: dict com 'template' e/ou 'params' para customizar prompt
    """
    # Configuração de template padrão se não fornecida
    if template_config is None:
        template_config = {'template': 'default', 'params': None}
    
    if ai_model == 'claude' and CLAUDE_AVAILABLE:
        return analyze_image_with_claude(image_path, api_key, claude_version, template_config)
    elif ai_model == 'gpt4' and GPT_AVAILABLE:
        return analyze_image_with_gpt4(image_path, api_key, template_config)
    elif ai_model == 'gemini' and GEMINI_AVAILABLE:
        return analyze_image_with_gemini(image_path, api_key, gemini_version, template_config)
    elif ai_model == 'deepseek' and DEEPSEEK_AVAILABLE:
        return analyze_image_with_deepseek(image_path, api_key, template_config)
    elif ai_model == 'qwen' and QWEN_AVAILABLE:
        return analyze_image_with_qwen(image_path, api_key, template_config)
    elif ai_model == 'huggingface' and HUGGINGFACE_AVAILABLE:
        return analyze_image_with_huggingface(image_path, api_key, template_config)
    else:
        # Fallback: tentar modelos gratuitos primeiro
        if DEEPSEEK_AVAILABLE and api_key:
            return analyze_image_with_deepseek(image_path, api_key, template_config)
        elif QWEN_AVAILABLE and api_key:
            return analyze_image_with_qwen(image_path, api_key, template_config)
        elif CLAUDE_AVAILABLE:
            return analyze_image_with_claude(image_path, api_key, claude_version, template_config)
        else:
            return {
                "especies": [
                    {
                        "apelido": "IA não disponível",
                        "cobertura": 0,
                        "altura": 0,
                        "forma_vida": "-",
                        "erro": f"IA {ai_model} não está disponível. Instale: pip install openai requests"
                    }
                ]
            }

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/upload', methods=['POST'])
def upload_images():
    """Recebe upload de múltiplas imagens"""
    if 'images' not in request.files:
        return jsonify({'error': 'Nenhuma imagem enviada'}), 400

    parcela = request.form.get('parcela', 'Parcela_1')
    files = request.files.getlist('images')

    if not files or len(files) == 0:
        return jsonify({'error': 'Nenhuma imagem selecionada'}), 400

    uploaded_files = []

    # Criar diretório para a parcela
    parcela_dir = os.path.join(app.config['UPLOAD_FOLDER'], parcela)
    os.makedirs(parcela_dir, exist_ok=True)

    for idx, file in enumerate(files, 1):
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(parcela_dir, filename)
            file.save(filepath)
            uploaded_files.append({
                'filename': filename,
                'path': filepath,
                'subparcela': idx
            })

    # Inicializar dados da parcela
    if parcela not in analysis_data['parcelas']:
        analysis_data['parcelas'][parcela] = {
            'images': [],
            'subparcelas': {}
        }

    analysis_data['parcelas'][parcela]['images'] = uploaded_files

    return jsonify({
        'success': True,
        'message': f'{len(uploaded_files)} imagens enviadas com sucesso',
        'parcela': parcela,
        'files': uploaded_files
    })

@app.route('/api/ai/available', methods=['GET'])
def get_available_ais():
    """Retorna lista de IAs disponíveis"""
    ais = []

    # Modelos Premium
    if CLAUDE_AVAILABLE:
        ais.append({
            'id': 'claude',
            'name': 'Claude',
            'provider': 'Anthropic',
            'tier': 'premium',
            'available': True
        })

    if GPT_AVAILABLE:
        ais.append({
            'id': 'gpt4',
            'name': 'GPT-4 Vision',
            'provider': 'OpenAI',
            'tier': 'premium',
            'available': True
        })

    if GEMINI_AVAILABLE:
        ais.append({
            'id': 'gemini',
            'name': 'Gemini',
            'provider': 'Google',
            'tier': 'free',
            'available': True
        })

    # Modelos Gratuitos/Open Source
    # NOTA: DeepSeek Chat não suporta análise de imagens, apenas texto
    # Removido temporariamente até que lancem modelo com visão

    if QWEN_AVAILABLE:
        ais.append({
            'id': 'qwen',
            'name': 'Qwen VL Max',
            'provider': 'Alibaba (Grátis)',
            'tier': 'free',
            'available': True
        })

    if HUGGINGFACE_AVAILABLE:
        ais.append({
            'id': 'huggingface',
            'name': 'HuggingFace LLaVA',
            'provider': 'HuggingFace (Grátis)',
            'tier': 'free',
            'available': True
        })

    return jsonify({
        'ais': ais,
        'default': app.config['DEFAULT_AI']  # Retornar AI padrão
    })

@app.route('/api/templates', methods=['GET'])
def get_templates():
    """Retorna lista de templates de prompt disponíveis"""
    templates = get_template_list()
    return jsonify({'templates': templates})

@app.route('/api/templates/<template_id>', methods=['GET'])
def get_template_detail(template_id):
    """Retorna detalhes de um template específico"""
    if template_id not in PROMPT_TEMPLATES:
        return jsonify({'error': 'Template não encontrado'}), 404
    
    template = PROMPT_TEMPLATES[template_id]
    return jsonify({
        'id': template_id,
        'name': template['name'],
        'description': template['description'],
        'objective': template['objective'],
        'params': template['params']
    })

@app.route('/api/templates/preview', methods=['POST'])
def preview_prompt():
    """Gera preview do prompt com parâmetros fornecidos"""
    data = request.get_json()
    template_name = data.get('template', 'default')
    custom_params = data.get('params', {})
    
    try:
        prompt = build_prompt(template_name, custom_params)
        return jsonify({
            'success': True,
            'prompt': prompt,
            'length': len(prompt)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/templates/custom/save', methods=['POST'])
def save_custom_template():
    """Salva um template customizado localmente"""
    try:
        data = request.get_json()
        template_name = data.get('name')
        template_params = data.get('params', {})
        template_prompt = data.get('prompt', '')
        
        if not template_name:
            return jsonify({'error': 'Nome do template é obrigatório'}), 400
        
        # Criar diretório para templates customizados
        custom_templates_dir = os.path.join(os.path.dirname(__file__), 'custom_templates')
        os.makedirs(custom_templates_dir, exist_ok=True)
        
        # Sanitizar nome do arquivo
        safe_name = "".join(c for c in template_name if c.isalnum() or c in (' ', '-', '_')).strip()
        template_file = os.path.join(custom_templates_dir, f"{safe_name}.json")
        
        # Preparar dados do template
        template_data = {
            'name': template_name,
            'params': template_params,
            'prompt': template_prompt,
            'created_at': datetime.now().isoformat(),
            'version': '1.0'
        }
        
        # Salvar como JSON
        with open(template_file, 'w', encoding='utf-8') as f:
            json.dump(template_data, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            'success': True,
            'message': f'Template "{template_name}" salvo com sucesso',
            'filename': f"{safe_name}.json"
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/templates/custom/list', methods=['GET'])
def list_custom_templates():
    """Lista todos os templates customizados salvos"""
    try:
        custom_templates_dir = os.path.join(os.path.dirname(__file__), 'custom_templates')
        
        if not os.path.exists(custom_templates_dir):
            return jsonify({'templates': []})
        
        templates = []
        for filename in os.listdir(custom_templates_dir):
            if filename.endswith('.json'):
                filepath = os.path.join(custom_templates_dir, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        template_data = json.load(f)
                        templates.append({
                            'filename': filename,
                            'name': template_data.get('name', filename),
                            'description': template_data.get('description', ''),
                            'created_at': template_data.get('created_at'),
                            'params': template_data.get('params', {}),
                            'params_count': len(template_data.get('params', {}))
                        })
                except Exception as e:
                    print(f"Erro ao ler template {filename}: {e}")
                    continue
        
        # Ordenar por data de criação (mais recente primeiro)
        templates.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        
        return jsonify({'templates': templates})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/templates/custom/load/<filename>', methods=['GET'])
def load_custom_template(filename):
    """Carrega um template customizado específico"""
    try:
        custom_templates_dir = os.path.join(os.path.dirname(__file__), 'custom_templates')
        filepath = os.path.join(custom_templates_dir, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'Template não encontrado'}), 404
        
        with open(filepath, 'r', encoding='utf-8') as f:
            template_data = json.load(f)
        
        return jsonify({
            'success': True,
            'template': template_data
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/templates/custom/delete/<filename>', methods=['DELETE'])
def delete_custom_template(filename):
    """Deleta um template customizado"""
    try:
        custom_templates_dir = os.path.join(os.path.dirname(__file__), 'custom_templates')
        filepath = os.path.join(custom_templates_dir, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'Template não encontrado'}), 404
        
        os.remove(filepath)
        
        return jsonify({
            'success': True,
            'message': 'Template deletado com sucesso'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/analyze/<parcela>', methods=['POST'])
def analyze_parcela(parcela):
    """Analisa todas as imagens de uma parcela com progresso em tempo real"""
    print(f"\n=== Iniciando análise da parcela: {parcela} ===")

    if parcela not in analysis_data['parcelas']:
        print(f"ERRO: Parcela {parcela} não encontrada")
        return jsonify({'error': 'Parcela não encontrada'}), 404

    # Obter modelo de IA selecionado e configuração de template
    data = request.get_json() or {}
    ai_model = data.get('ai_model', app.config['DEFAULT_AI'])
    template_config = data.get('template_config', {'template': 'default', 'params': None})
    
    print(f"Modelo de IA selecionado: {ai_model}")
    print(f"Template de prompt: {template_config.get('template', 'default')}")

    # Obter versão específica do Gemini (se aplicável)
    gemini_version = request.headers.get('X-Gemini-Version', 'gemini-flash-latest')
    if ai_model == 'gemini':
        print(f"Versão do Gemini selecionada: {gemini_version}")
    
    # Obter versão específica do Claude (se aplicável)
    claude_version = request.headers.get('X-Claude-Version', 'claude-sonnet-4-5-20250929')
    if ai_model == 'claude':
        print(f"Versão do Claude selecionada: {claude_version}")

    # Obter API keys dos headers (decodificar de Base64)
    api_key = None
    if ai_model == 'claude':
        api_key = decode_api_key(request.headers.get('X-API-Key-Claude'))
        print(f"🔑 Claude API Key recebida: {api_key[:15] if api_key else 'VAZIA'}... (tamanho: {len(api_key) if api_key else 0})")
    elif ai_model == 'gpt4':
        api_key = decode_api_key(request.headers.get('X-API-Key-GPT4'))
    elif ai_model == 'gemini':
        api_key = decode_api_key(request.headers.get('X-API-Key-Gemini'))
        print(f"🔑 Gemini API Key recebida: {api_key[:15] if api_key else 'VAZIA'}... (tamanho: {len(api_key) if api_key else 0})")
    elif ai_model == 'deepseek':
        api_key = decode_api_key(request.headers.get('X-API-Key-DeepSeek'))
    elif ai_model == 'qwen':
        api_key = decode_api_key(request.headers.get('X-API-Key-Qwen'))
    elif ai_model == 'huggingface':
        api_key = decode_api_key(request.headers.get('X-API-Key-HuggingFace'))

    if not api_key:
        print(f"ERRO: API key não fornecida para {ai_model}")
        return jsonify({'error': f'API key não configurada para {ai_model}'}), 400

    print(f"API key presente: {api_key[:10]}..." if api_key else "Nenhuma")

    def generate():
        """Generator para enviar eventos de progresso"""
        parcela_info = analysis_data['parcelas'][parcela]
        results = []
        total_images = len(parcela_info['images'])

        # Evento inicial
        yield f"data: {json.dumps({'type': 'start', 'total': total_images})}\n\n"
        sys.stdout.flush()  # Forçar envio imediato

        for idx, img_info in enumerate(parcela_info['images'], 1):
            # Verificar se img_info é um dicionário ou string
            if isinstance(img_info, str):
                # Se for string, é apenas o filename - precisa reconstruir a estrutura
                print(f"⚠️ AVISO: img_info é string: {img_info}")
                filename = img_info
                parcela_dir = os.path.join(app.config['UPLOAD_FOLDER'], parcela)
                filepath = os.path.join(parcela_dir, filename)
                subparcela = idx
                img_info = {
                    'filename': filename,
                    'path': filepath,
                    'subparcela': idx
                }
            else:
                subparcela = img_info['subparcela']
                filepath = img_info['path']
            
            # Evento de progresso: iniciando análise
            percentage = int((idx - 1) / total_images * 100)
            yield f"data: {json.dumps({'type': 'progress', 'current': idx-1, 'total': total_images, 'percentage': percentage, 'subparcela': subparcela, 'status': 'analyzing'})}\n\n"
            sys.stdout.flush()  # Forçar envio imediato

            print(f"\nAnalisando subparcela {subparcela} ({idx}/{total_images})")
            print(f"Arquivo: {filepath}")

            try:
                # Obter apelidos existentes para padronização (aninhado por parcela)
                existing_apelidos = list(analysis_data['especies_unificadas'].get(parcela, {}).keys())
                
                # Carregar espécies de referência
                reference_apelidos = []
                try:
                    ref_file = os.path.join(os.path.dirname(__file__), 'reference_species.json')
                    if os.path.exists(ref_file):
                        with open(ref_file, 'r', encoding='utf-8') as f:
                            ref_data = json.load(f)
                            
                            # Suportar ambos os formatos: {species: [...]} ou [...]
                            if isinstance(ref_data, dict):
                                ref_species = ref_data.get('species', [])
                            elif isinstance(ref_data, list):
                                ref_species = ref_data
                            else:
                                ref_species = []
                            
                            # Extrair apelidos
                            for sp in ref_species:
                                if isinstance(sp, dict) and sp.get('apelido'):
                                    reference_apelidos.append(sp['apelido'])
                                elif isinstance(sp, str):
                                    reference_apelidos.append(sp)
                            
                            if reference_apelidos:
                                print(f"📚 Espécies de referência carregadas: {reference_apelidos}")
                except Exception as e:
                    print(f"⚠️ Erro ao carregar espécies de referência: {e}")
                
                # Combinar apelidos existentes + referências (sem duplicatas)
                all_apelidos = list(set(existing_apelidos + reference_apelidos))
                
                if len(all_apelidos) > 0:
                    print(f"Apelidos para padronização: {all_apelidos}")
                    # Adicionar aos parâmetros do template
                    if template_config.get('params') is None:
                        template_config['params'] = {}
                    template_config['params']['existing_species'] = all_apelidos
                
                # Analisar imagem com IA selecionada (com retry se retornar vazio)
                max_retries = 2
                analysis = None
                especies_validas = []
                
                for retry in range(max_retries + 1):
                    if retry > 0:
                        print(f"⚠️ Tentativa {retry + 1}/{max_retries + 1} - análise anterior retornou vazia")
                        yield f"data: {json.dumps({'type': 'progress', 'current': idx-1, 'total': total_images, 'percentage': percentage, 'subparcela': subparcela, 'status': 'retrying', 'retry': retry})}\n\n"
                        sys.stdout.flush()  # Forçar envio imediato

                    if ai_model == 'gemini':
                        analysis = analyze_image_with_ai(filepath, ai_model, api_key, gemini_version, None, template_config)
                    elif ai_model == 'claude':
                        analysis = analyze_image_with_ai(filepath, ai_model, api_key, None, claude_version, template_config)
                    else:
                        analysis = analyze_image_with_ai(filepath, ai_model, api_key, None, None, template_config)
                    
                    # Validar se retornou espécies válidas
                    especies_validas = []
                    for esp in analysis.get('especies', []):
                        apelido = esp.get('apelido', '')
                        
                        # FILTRAR mensagens de erro
                        erro_keywords = [
                            'erro', 'error', 'falha', 'limite', 'atingido', 
                            'aguarde', 'não disponível', 'timeout', 'quota',
                            'exceeded', 'rate limit', 'api key', 'invalid'
                        ]
                        
                        apelido_lower = apelido.lower()
                        is_error = any(keyword in apelido_lower for keyword in erro_keywords)
                        
                        if not is_error and 'erro' not in esp:
                            especies_validas.append(esp)
                    
                    # Se encontrou espécies válidas, sair do loop
                    if len(especies_validas) > 0:
                        print(f"✓ Análise válida: {len(especies_validas)} espécies detectadas")
                        break
                    
                    # Se foi a última tentativa e ainda está vazio
                    if retry == max_retries:
                        print(f"⚠️ AVISO: Após {max_retries + 1} tentativas, nenhuma espécie válida detectada")
                        # Criar espécie placeholder para indicar área sem vegetação detectável
                        especies_validas = [{
                            'apelido': 'Vegetação Não Detectada',
                            'genero': '',
                            'familia': '',
                            'observacoes': 'IA não conseguiu identificar vegetação clara nesta subparcela após múltiplas tentativas. Pode ser área com vegetação muito esparsa, solo exposto, ou imagem de baixa qualidade. Revise manualmente.',
                            'cobertura': 100,
                            'altura': 0,
                            'forma_vida': '-'
                        }]
                        analysis['especies'] = especies_validas
                
                num_especies = len(especies_validas)
                
                # Evento de progresso: análise concluída
                percentage = int(idx / total_images * 100)
                yield f"data: {json.dumps({'type': 'progress', 'current': idx, 'total': total_images, 'percentage': percentage, 'subparcela': subparcela, 'status': 'completed', 'especies_count': num_especies})}\n\n"
                sys.stdout.flush()  # Forçar envio imediato

            except Exception as e:
                error_msg = str(e)
                print(f"ERRO na análise da subparcela {subparcela}: {error_msg}")
                
                # Evento de erro
                yield f"data: {json.dumps({'type': 'error', 'subparcela': subparcela, 'error': error_msg[:100]})}\n\n"
                sys.stdout.flush()  # Forçar envio imediato

                analysis = {
                    "especies": [{
                        "apelido": "Erro na análise",
                        "cobertura": 0,
                        "altura": 0,
                        "forma_vida": "-",
                        "erro": error_msg
                    }]
                }

            # Processar resultados
            especies_encontradas = []
            for esp_idx, esp in enumerate(analysis.get('especies', []), 1):
                apelido = esp['apelido']

                # FILTRAR mensagens de erro que não são espécies
                erro_keywords = [
                    'erro', 'error', 'falha', 'limite', 'atingido', 
                    'aguarde', 'não disponível', 'timeout', 'quota',
                    'exceeded', 'rate limit', 'api key', 'invalid'
                ]
                
                apelido_lower = apelido.lower()
                is_error = any(keyword in apelido_lower for keyword in erro_keywords)
                
                if is_error or 'erro' in esp:
                    print(f"⚠️  IGNORANDO mensagem de erro: {apelido}")
                    continue

                # Garantir que especies_unificadas está aninhado por parcela
                if parcela not in analysis_data['especies_unificadas']:
                    analysis_data['especies_unificadas'][parcela] = {}
                
                # Adicionar à lista unificada de espécies
                if apelido not in analysis_data['especies_unificadas'][parcela]:
                    analysis_data['especies_unificadas'][parcela][apelido] = {
                        'apelido_original': apelido,
                        'apelido_usuario': apelido,
                        'genero': esp.get('genero', ''),
                        'especie': '',
                        'familia': esp.get('familia', ''),
                        'observacoes': esp.get('observacoes', ''),
                        'ocorrencias': 0
                    }
                else:
                    # Atualizar genero/familia/observacoes se vieram preenchidos da IA
                    if esp.get('genero') and not analysis_data['especies_unificadas'][parcela][apelido]['genero']:
                        analysis_data['especies_unificadas'][parcela][apelido]['genero'] = esp.get('genero')
                    if esp.get('familia') and not analysis_data['especies_unificadas'][parcela][apelido]['familia']:
                        analysis_data['especies_unificadas'][parcela][apelido]['familia'] = esp.get('familia')
                    if esp.get('observacoes') and not analysis_data['especies_unificadas'][parcela][apelido]['observacoes']:
                        analysis_data['especies_unificadas'][parcela][apelido]['observacoes'] = esp.get('observacoes')

                analysis_data['especies_unificadas'][parcela][apelido]['ocorrencias'] += 1

                especies_encontradas.append({
                    'indice': esp_idx,
                    'apelido': apelido,
                    'genero': esp.get('genero', ''),
                    'familia': esp.get('familia', ''),
                    'observacoes': esp.get('observacoes', ''),
                    'cobertura': esp['cobertura'],
                    'altura': esp['altura'],
                    'forma_vida': esp['forma_vida']
                })

            parcela_info['subparcelas'][subparcela] = {
                'image': img_info['filename'],
                'especies': especies_encontradas
            }

            results.append({
                'subparcela': subparcela,
                'image': img_info['filename'],
                'especies': especies_encontradas
            })
            
            # 📊 Enviar resumo acumulativo após processar cada subparcela
            total_especies_unicas = len(analysis_data['especies_unificadas'].get(parcela, {}))
            especies_resumo = []
            for apelido, info in analysis_data['especies_unificadas'].get(parcela, {}).items():
                especies_resumo.append({
                    'apelido': apelido,
                    'ocorrencias': info.get('ocorrencias', 0)
                })
            
            # Ordenar por ocorrências (mais frequentes primeiro)
            especies_resumo.sort(key=lambda x: x['ocorrencias'], reverse=True)
            
            percentage_update = int((idx + 1) / total_images * 100)
            yield f"data: {json.dumps({'type': 'progress', 'current': idx + 1, 'total': total_images, 'percentage': percentage_update, 'subparcela': subparcela, 'status': 'summary', 'total_especies_unicas': total_especies_unicas, 'especies_resumo': especies_resumo[:10]})}\n\n"
            sys.stdout.flush()  # Forçar envio imediato

        # Evento final
        yield f"data: {json.dumps({'type': 'complete', 'success': True, 'parcela': parcela, 'results': results, 'especies_unificadas': analysis_data['especies_unificadas'].get(parcela, {})})}\n\n"
        sys.stdout.flush()  # Forçar envio final

    response = Response(stream_with_context(generate()), content_type='text/event-stream')
    response.headers['Cache-Control'] = 'no-cache'
    response.headers['X-Accel-Buffering'] = 'no'
    response.headers['Connection'] = 'keep-alive'
    return response


@app.route('/api/upload-additional-images', methods=['POST'])
def upload_additional_images():
    """Upload de novas imagens para adicionar a uma análise existente"""
    if 'images' not in request.files:
        return jsonify({'error': 'Nenhuma imagem enviada'}), 400

    parcela_nome = request.form.get('parcela_nome')
    if not parcela_nome:
        return jsonify({'error': 'Nome da parcela não fornecido'}), 400
    
    if parcela_nome not in analysis_data['parcelas']:
        return jsonify({'error': 'Parcela não encontrada'}), 404

    files = request.files.getlist('images')
    if not files or len(files) == 0:
        return jsonify({'error': 'Nenhuma imagem selecionada'}), 400

    # Obter número atual de subparcelas para continuar a numeração
    parcela_info = analysis_data['parcelas'][parcela_nome]
    current_subparcelas = len(parcela_info.get('subparcelas', {}))
    next_subparcela_id = current_subparcelas + 1

    # Diretório da parcela
    parcela_dir = os.path.join(app.config['UPLOAD_FOLDER'], parcela_nome)
    os.makedirs(parcela_dir, exist_ok=True)

    uploaded_files = []
    subparcela_ids = []

    for file in files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(parcela_dir, filename)
            file.save(filepath)
            
            subparcela_id = next_subparcela_id
            next_subparcela_id += 1
            
            uploaded_files.append({
                'filename': filename,
                'path': filepath,
                'subparcela': subparcela_id
            })
            subparcela_ids.append(subparcela_id)

    # Adicionar às imagens da parcela
    if 'images' not in parcela_info:
        parcela_info['images'] = []
    
    parcela_info['images'].extend(uploaded_files)

    print(f"✓ {len(uploaded_files)} novas imagens adicionadas à parcela {parcela_nome}")

    return jsonify({
        'success': True,
        'message': f'{len(uploaded_files)} imagens enviadas com sucesso',
        'parcela_nome': parcela_nome,
        'subparcela_ids': subparcela_ids,
        'files': uploaded_files
    })


@app.route('/api/analyze-additional-images', methods=['POST'])
def analyze_additional_images():
    """Analisa novas imagens adicionadas a uma parcela existente"""
    data = request.get_json()
    
    parcela_nome = data.get('parcela_nome')
    subparcela_ids = data.get('subparcela_ids', [])
    ai_model = data.get('ai_model', app.config['DEFAULT_AI'])
    api_key = data.get('api_key')
    prompt_config = data.get('prompt_config', {'template': 'default', 'params': None})
    
    if not parcela_nome or not subparcela_ids:
        return jsonify({'error': 'Parâmetros inválidos'}), 400
    
    if parcela_nome not in analysis_data['parcelas']:
        return jsonify({'error': 'Parcela não encontrada'}), 404
    
    if not api_key:
        return jsonify({'error': f'API key não fornecida para {ai_model}'}), 400

    parcela_info = analysis_data['parcelas'][parcela_nome]
    
    # Obter apelidos existentes para padronização (aninhado por parcela)
    existing_apelidos = list(analysis_data['especies_unificadas'].get(parcela_nome, {}).keys())
    
    # Carregar espécies de referência
    reference_apelidos = []
    try:
        ref_file = os.path.join(os.path.dirname(__file__), 'reference_species.json')
        if os.path.exists(ref_file):
            with open(ref_file, 'r', encoding='utf-8') as f:
                ref_data = json.load(f)
                
                # Suportar ambos os formatos: {species: [...]} ou [...]
                if isinstance(ref_data, dict):
                    ref_species = ref_data.get('species', [])
                elif isinstance(ref_data, list):
                    ref_species = ref_data
                else:
                    ref_species = []
                
                # Extrair apelidos
                for sp in ref_species:
                    if isinstance(sp, dict) and sp.get('apelido'):
                        reference_apelidos.append(sp['apelido'])
                    elif isinstance(sp, str):
                        reference_apelidos.append(sp)
                
                if reference_apelidos:
                    print(f"📚 Espécies de referência carregadas: {reference_apelidos}")
    except Exception as e:
        print(f"⚠️ Erro ao carregar espécies de referência: {e}")
    
    # Combinar apelidos existentes + referências (sem duplicatas)
    all_apelidos = list(set(existing_apelidos + reference_apelidos))
    
    if len(all_apelidos) > 0:
        if prompt_config.get('params') is None:
            prompt_config['params'] = {}
        prompt_config['params']['existing_species'] = all_apelidos
        print(f"Apelidos para padronização: {all_apelidos}")

    novas_subparcelas = []
    
    # Analisar apenas as novas subparcelas
    for img_info in parcela_info['images']:
        subparcela_id = img_info['subparcela']
        
        if subparcela_id not in subparcela_ids:
            continue
        
        filepath = img_info['path']
        
        print(f"\nAnalisando nova subparcela {subparcela_id}")
        print(f"Arquivo: {filepath}")
        
        try:
            # Analisar imagem
            analysis = analyze_image_with_ai(
                filepath, 
                ai_model=ai_model, 
                api_key=api_key,
                template_config=prompt_config
            )
            
            num_especies = len(analysis.get('especies', []))
            print(f"✓ Análise concluída: {num_especies} espécies encontradas")
            
        except Exception as e:
            error_msg = str(e)
            print(f"ERRO na análise da subparcela {subparcela_id}: {error_msg}")
            
            analysis = {
                "especies": [{
                    "apelido": "Erro na análise",
                    "cobertura": 0,
                    "altura": 0,
                    "forma_vida": "-",
                    "erro": error_msg
                }]
            }
        
        # Processar resultados (mesmo código da análise principal)
        especies_encontradas = []
        for esp_idx, esp in enumerate(analysis.get('especies', []), 1):
            apelido = esp['apelido']
            
            # Filtrar mensagens de erro
            erro_keywords = [
                'erro', 'error', 'falha', 'limite', 'atingido', 
                'aguarde', 'não disponível', 'timeout', 'quota',
                'exceeded', 'rate limit', 'api key', 'invalid'
            ]
            
            apelido_lower = apelido.lower()
            is_error = any(keyword in apelido_lower for keyword in erro_keywords)
            
            if is_error or 'erro' in esp:
                print(f"⚠️  IGNORANDO mensagem de erro: {apelido}")
                continue
            
            # Adicionar/atualizar espécie unificada (com parcela)
            if parcela_nome not in analysis_data['especies_unificadas']:
                analysis_data['especies_unificadas'][parcela_nome] = {}
                
            if apelido not in analysis_data['especies_unificadas'][parcela_nome]:
                analysis_data['especies_unificadas'][parcela_nome][apelido] = {
                    'apelido_original': apelido,
                    'apelido_usuario': apelido,
                    'genero': esp.get('genero', ''),
                    'especie': '',
                    'familia': esp.get('familia', ''),
                    'observacoes': esp.get('observacoes', ''),
                    'ocorrencias': 0
                }
            else:
                if esp.get('genero') and not analysis_data['especies_unificadas'][parcela_nome][apelido]['genero']:
                    analysis_data['especies_unificadas'][parcela_nome][apelido]['genero'] = esp.get('genero')
                if esp.get('familia') and not analysis_data['especies_unificadas'][parcela_nome][apelido]['familia']:
                    analysis_data['especies_unificadas'][parcela_nome][apelido]['familia'] = esp.get('familia')
                if esp.get('observacoes') and not analysis_data['especies_unificadas'][parcela_nome][apelido]['observacoes']:
                    analysis_data['especies_unificadas'][parcela_nome][apelido]['observacoes'] = esp.get('observacoes')
            
            analysis_data['especies_unificadas'][parcela_nome][apelido]['ocorrencias'] += 1
            
            especies_encontradas.append({
                'indice': esp_idx,
                'apelido': apelido,
                'genero': esp.get('genero', ''),
                'familia': esp.get('familia', ''),
                'observacoes': esp.get('observacoes', ''),
                'cobertura': esp['cobertura'],
                'altura': esp['altura'],
                'forma_vida': esp['forma_vida']
            })
        
        # Adicionar subparcela aos resultados
        parcela_info['subparcelas'][subparcela_id] = {
            'image': img_info['filename'],
            'image_path': filepath,
            'especies': especies_encontradas
        }
        
        novas_subparcelas.append({
            'subparcela_id': subparcela_id,
            'image_path': filepath,
            'especies': especies_encontradas,
            'analise_completa': True
        })
    
    print(f"✓ {len(novas_subparcelas)} novas subparcelas analisadas")
    
    # Retornar espécies da parcela para compatibilidade
    especies_retorno = analysis_data['especies_unificadas'].get(parcela_nome, {})
    
    return jsonify({
        'success': True,
        'message': f'{len(novas_subparcelas)} subparcelas analisadas com sucesso',
        'novas_subparcelas': novas_subparcelas,
        'especies_atualizadas': especies_retorno
    })


@app.route('/api/especies', methods=['GET'])
def get_especies():
    """Retorna lista unificada de espécies"""
    # Retornar espécies da primeira/única parcela para compatibilidade
    # (sistema atualmente focado em single-parcela)
    especies = {}
    if analysis_data['especies_unificadas']:
        # Pegar primeira parcela
        first_parcela = next(iter(analysis_data['especies_unificadas'].values()))
        especies = first_parcela if isinstance(first_parcela, dict) else {}
    
    return jsonify({
        'especies': especies
    })

@app.route('/api/especies/<apelido_original>', methods=['PUT'])
def update_especie(apelido_original):
    """Atualiza informações de uma espécie"""
    # Procurar espécie em todas as parcelas
    especie_encontrada = False
    parcela_da_especie = None
    
    for parcela_nome, especies_parcela in analysis_data['especies_unificadas'].items():
        if apelido_original in especies_parcela:
            especie_encontrada = True
            parcela_da_especie = parcela_nome
            break
    
    if not especie_encontrada:
        return jsonify({'error': 'Espécie não encontrada'}), 404

    data = request.json
    especie = analysis_data['especies_unificadas'][parcela_da_especie][apelido_original]

    # Atualizar dados na lista unificada
    if 'apelido_usuario' in data:
        especie['apelido_usuario'] = data['apelido_usuario']
    if 'genero' in data:
        especie['genero'] = data['genero']
    if 'especie' in data:
        especie['especie'] = data['especie']
    if 'familia' in data:
        especie['familia'] = data['familia']
    if 'link_fotos' in data:
        especie['link_fotos'] = data['link_fotos']
    
    # CRÍTICO: Propagar mudanças para todas as subparcelas que contenham esta espécie
    for parcela_nome, parcela in analysis_data['parcelas'].items():
        for subparcela_num, subparcela in parcela.get('subparcelas', {}).items():
            for esp in subparcela.get('especies', []):
                # Se o apelido corresponde (pode ser apelido_original ou apelido_usuario antigo)
                if esp.get('apelido') == apelido_original:
                    # 🔧 FIX: Atualizar TAMBÉM o apelido exibido (não apenas taxonomia)
                    if 'apelido_usuario' in data:
                        esp['apelido'] = data['apelido_usuario']
                    # Atualizar informações taxonômicas nas subparcelas
                    if 'genero' in data:
                        esp['genero'] = data['genero']
                    if 'especie' in data:
                        esp['especie'] = data['especie']
                    if 'familia' in data:
                        esp['familia'] = data['familia']
                    if 'link_fotos' in data:
                        esp['link_fotos'] = data['link_fotos']

    print(f"✓ Espécie '{apelido_original}' atualizada em especies_unificadas e propagada para subparcelas (incluindo apelido exibido e link fotos)")

    return jsonify({
        'success': True,
        'especie': especie
    })

@app.route('/api/especies/merge', methods=['POST'])
def merge_especies():
    """Unifica múltiplas espécies em uma única espécie"""
    data = request.json
    especies_origem = data.get('especies_origem', [])  # Lista de apelidos a unificar
    novo_apelido = data.get('novo_apelido', '')

    if len(especies_origem) < 2:
        return jsonify({'error': 'Selecione pelo menos 2 espécies para unificar'}), 400

    if not novo_apelido:
        return jsonify({'error': 'Forneça um novo apelido para a espécie unificada'}), 400

    # Criar nova espécie unificada
    nova_especie = {
        'apelido_original': novo_apelido,
        'apelido_usuario': novo_apelido,
        'genero': data.get('genero', ''),
        'especie': data.get('especie', ''),
        'familia': data.get('familia', ''),
        'ocorrencias': 0
    }

    # Atualizar todas as ocorrências nas subparcelas
    for parcela_nome, parcela in analysis_data['parcelas'].items():
        for subparcela_num, subparcela in parcela.get('subparcelas', {}).items():
            especies_atualizadas = []
            especies_a_mesclar = []

            for esp in subparcela['especies']:
                if esp['apelido'] in especies_origem:
                    especies_a_mesclar.append(esp)
                else:
                    especies_atualizadas.append(esp)

            # Se houver espécies a mesclar nesta subparcela
            if especies_a_mesclar:
                # Somar coberturas e calcular média de alturas
                cobertura_total = sum(e['cobertura'] for e in especies_a_mesclar)
                altura_media = sum(e['altura'] * e['cobertura'] for e in especies_a_mesclar) / cobertura_total if cobertura_total > 0 else 0

                # Pegar forma de vida da primeira espécie (ou a mais comum)
                forma_vida = especies_a_mesclar[0]['forma_vida']

                # Adicionar espécie mesclada
                especies_atualizadas.append({
                    'indice': len(especies_atualizadas) + 1,
                    'apelido': novo_apelido,
                    'cobertura': round(cobertura_total, 1),
                    'altura': round(altura_media, 1),
                    'forma_vida': forma_vida
                })

                nova_especie['ocorrencias'] += 1

            # Reindexar espécies
            for idx, esp in enumerate(especies_atualizadas, 1):
                esp['indice'] = idx

            subparcela['especies'] = especies_atualizadas

    # Remover espécies antigas da lista unificada (em todas as parcelas)
    for parcela_nome, especies_parcela in analysis_data['especies_unificadas'].items():
        for apelido in especies_origem:
            if apelido in especies_parcela:
                del especies_parcela[apelido]

    # Adicionar nova espécie na primeira parcela (compatibilidade)
    if analysis_data['especies_unificadas']:
        first_parcela = next(iter(analysis_data['especies_unificadas'].keys()))
        analysis_data['especies_unificadas'][first_parcela][novo_apelido] = nova_especie

    return jsonify({
        'success': True,
        'nova_especie': nova_especie,
        'message': f'{len(especies_origem)} espécies unificadas em "{novo_apelido}"'
    })

@app.route('/api/especies/split', methods=['POST'])
def split_especie():
    """Subdivide uma espécie em múltiplas espécies dentro de uma subparcela"""
    data = request.json
    parcela_nome = data.get('parcela')
    subparcela_num = data.get('subparcela')
    apelido_original = data.get('apelido_original')
    novas_especies = data.get('novas_especies', [])  # Lista de {apelido, cobertura, altura, forma_vida}

    if not all([parcela_nome, subparcela_num, apelido_original, novas_especies]):
        return jsonify({'error': 'Dados incompletos'}), 400

    if parcela_nome not in analysis_data['parcelas']:
        return jsonify({'error': 'Parcela não encontrada'}), 404

    parcela = analysis_data['parcelas'][parcela_nome]
    if subparcela_num not in parcela.get('subparcelas', {}):
        return jsonify({'error': 'Subparcela não encontrada'}), 404

    subparcela = parcela['subparcelas'][subparcela_num]

    # Remover espécie original
    especies_atualizadas = [e for e in subparcela['especies'] if e['apelido'] != apelido_original]

    # Adicionar novas espécies
    for nova_esp in novas_especies:
        apelido = nova_esp['apelido']

        # Garantir que especies_unificadas está aninhado por parcela
        if parcela_nome not in analysis_data['especies_unificadas']:
            analysis_data['especies_unificadas'][parcela_nome] = {}

        # Adicionar à lista unificada se não existir
        if apelido not in analysis_data['especies_unificadas'][parcela_nome]:
            analysis_data['especies_unificadas'][parcela_nome][apelido] = {
                'apelido_original': apelido,
                'apelido_usuario': apelido,
                'genero': nova_esp.get('genero', ''),
                'especie': nova_esp.get('especie', ''),
                'familia': nova_esp.get('familia', ''),
                'ocorrencias': 0
            }

        analysis_data['especies_unificadas'][parcela_nome][apelido]['ocorrencias'] += 1

        especies_atualizadas.append({
            'indice': len(especies_atualizadas) + 1,
            'apelido': apelido,
            'cobertura': nova_esp['cobertura'],
            'altura': nova_esp['altura'],
            'forma_vida': nova_esp['forma_vida']
        })

    # Reindexar
    for idx, esp in enumerate(especies_atualizadas, 1):
        esp['indice'] = idx

    subparcela['especies'] = especies_atualizadas

    # Atualizar contagem de ocorrências da espécie original
    if parcela_nome in analysis_data['especies_unificadas'] and apelido_original in analysis_data['especies_unificadas'][parcela_nome]:
        analysis_data['especies_unificadas'][parcela_nome][apelido_original]['ocorrencias'] -= 1
        if analysis_data['especies_unificadas'][parcela_nome][apelido_original]['ocorrencias'] <= 0:
            del analysis_data['especies_unificadas'][parcela_nome][apelido_original]

    return jsonify({
        'success': True,
        'message': f'Espécie "{apelido_original}" subdividida em {len(novas_especies)} espécies'
    })

@app.route('/api/especies/add', methods=['POST'])
def add_especie():
    """Adiciona uma nova espécie em uma subparcela"""
    data = request.json
    parcela_nome = data.get('parcela')
    subparcela_num = data.get('subparcela')
    nova_especie = data.get('especie')  # {apelido, cobertura, altura, forma_vida}

    if not all([parcela_nome, subparcela_num, nova_especie]):
        return jsonify({'error': 'Dados incompletos'}), 400

    if parcela_nome not in analysis_data['parcelas']:
        return jsonify({'error': 'Parcela não encontrada'}), 404

    parcela = analysis_data['parcelas'][parcela_nome]
    if subparcela_num not in parcela.get('subparcelas', {}):
        return jsonify({'error': 'Subparcela não encontrada'}), 404

    subparcela = parcela['subparcelas'][subparcela_num]
    apelido = nova_especie['apelido']

    # Garantir que especies_unificadas está aninhado por parcela
    if parcela_nome not in analysis_data['especies_unificadas']:
        analysis_data['especies_unificadas'][parcela_nome] = {}

    # Adicionar à lista unificada se não existir
    if apelido not in analysis_data['especies_unificadas'][parcela_nome]:
        analysis_data['especies_unificadas'][parcela_nome][apelido] = {
            'apelido_original': apelido,
            'apelido_usuario': apelido,
            'genero': nova_especie.get('genero', ''),
            'especie': nova_especie.get('especie', ''),
            'familia': nova_especie.get('familia', ''),
            'link_fotos': nova_especie.get('link_fotos', ''),
            'ocorrencias': 0
        }

    analysis_data['especies_unificadas'][parcela_nome][apelido]['ocorrencias'] += 1

    # Adicionar espécie à subparcela
    subparcela['especies'].append({
        'indice': len(subparcela['especies']) + 1,
        'apelido': apelido,
        'cobertura': nova_especie['cobertura'],
        'altura': nova_especie['altura'],
        'forma_vida': nova_especie['forma_vida'],
        'genero': nova_especie.get('genero', ''),
        'especie': nova_especie.get('especie', ''),
        'familia': nova_especie.get('familia', ''),
        'link_fotos': nova_especie.get('link_fotos', '')
    })

    return jsonify({
        'success': True,
        'message': f'Espécie "{apelido}" adicionada com sucesso'
    })

@app.route('/api/especies/remove', methods=['POST'])
def remove_especie():
    """Remove uma espécie de uma subparcela específica"""
    data = request.json
    parcela_nome = data.get('parcela')
    subparcela_num = data.get('subparcela')
    apelido = data.get('apelido')

    if not all([parcela_nome, subparcela_num, apelido]):
        return jsonify({'error': 'Dados incompletos'}), 400

    if parcela_nome not in analysis_data['parcelas']:
        return jsonify({'error': 'Parcela não encontrada'}), 404

    parcela = analysis_data['parcelas'][parcela_nome]
    if subparcela_num not in parcela.get('subparcelas', {}):
        return jsonify({'error': 'Subparcela não encontrada'}), 404

    subparcela = parcela['subparcelas'][subparcela_num]

    # Remover espécie
    especies_atualizadas = [e for e in subparcela['especies'] if e['apelido'] != apelido]

    if len(especies_atualizadas) == len(subparcela['especies']):
        return jsonify({'error': 'Espécie não encontrada nesta subparcela'}), 404

    # Reindexar
    for idx, esp in enumerate(especies_atualizadas, 1):
        esp['indice'] = idx

    subparcela['especies'] = especies_atualizadas

    # Atualizar contagem de ocorrências
    if parcela_nome in analysis_data['especies_unificadas'] and apelido in analysis_data['especies_unificadas'][parcela_nome]:
        analysis_data['especies_unificadas'][parcela_nome][apelido]['ocorrencias'] -= 1
        if analysis_data['especies_unificadas'][parcela_nome][apelido]['ocorrencias'] <= 0:
            del analysis_data['especies_unificadas'][parcela_nome][apelido]

    return jsonify({
        'success': True,
        'message': f'Espécie "{apelido}" removida com sucesso'
    })

@app.route('/api/especies/<parcela>/<int:subparcela>/<apelido>', methods=['PUT'])
def update_especie_subparcela(parcela, subparcela, apelido):
    """Atualiza cobertura e altura de uma espécie em uma subparcela específica"""
    data = request.json

    if parcela not in analysis_data['parcelas']:
        return jsonify({'error': 'Parcela não encontrada'}), 404

    parcela_data = analysis_data['parcelas'][parcela]
    if subparcela not in parcela_data.get('subparcelas', {}):
        return jsonify({'error': 'Subparcela não encontrada'}), 404

    subparcela_data = parcela_data['subparcelas'][subparcela]

    # Encontrar e atualizar espécie
    especie_encontrada = False
    for esp in subparcela_data['especies']:
        if esp['apelido'] == apelido:
            # Atualizar todos os campos enviados
            if 'familia' in data:
                esp['familia'] = data['familia']
            if 'genero' in data:
                esp['genero'] = data['genero']
            if 'especie' in data:
                esp['especie'] = data['especie']
            if 'observacoes' in data:
                esp['observacoes'] = data['observacoes']
            if 'cobertura' in data:
                esp['cobertura'] = data['cobertura']
            if 'altura' in data:
                esp['altura'] = data['altura']
            if 'forma_vida' in data:
                esp['forma_vida'] = data['forma_vida']
            if 'link_fotos' in data:
                esp['link_fotos'] = data['link_fotos']
            especie_encontrada = True
            break

    if not especie_encontrada:
        return jsonify({'error': 'Espécie não encontrada nesta subparcela'}), 404

    # Atualizar especies_unificadas se campos taxonômicos foram alterados
    if parcela in analysis_data['especies_unificadas'] and apelido in analysis_data['especies_unificadas'][parcela]:
        especie_unificada = analysis_data['especies_unificadas'][parcela][apelido]
        if 'familia' in data:
            especie_unificada['familia'] = data['familia']
        if 'genero' in data:
            especie_unificada['genero'] = data['genero']
        if 'especie' in data:
            especie_unificada['especie'] = data['especie']
        if 'link_fotos' in data:
            especie_unificada['link_fotos'] = data['link_fotos']
        print(f"✓ Espécie '{apelido}' também atualizada em especies_unificadas (incluindo link fotos)")

    return jsonify({
        'success': True,
        'message': 'Espécie atualizada com sucesso'
    })

# NOVAS ROTAS para edição inline no modal
@app.route('/api/parcela/<parcela>/subparcela/<int:subparcela>/especie/<int:index>', methods=['PATCH'])
def update_especie_by_index(parcela, subparcela, index):
    """Atualiza uma espécie específica pelo índice na subparcela"""
    data = request.json

    if parcela not in analysis_data['parcelas']:
        return jsonify({'error': 'Parcela não encontrada'}), 404

    parcela_data = analysis_data['parcelas'][parcela]
    if subparcela not in parcela_data.get('subparcelas', {}):
        return jsonify({'error': 'Subparcela não encontrada'}), 404

    subparcela_data = parcela_data['subparcelas'][subparcela]
    
    if index < 0 or index >= len(subparcela_data['especies']):
        return jsonify({'error': 'Índice de espécie inválido'}), 404

    especie = subparcela_data['especies'][index]
    
    # Atualizar campos (incluindo taxonomia e observações)
    for field in ['apelido', 'genero', 'familia', 'observacoes', 'cobertura', 'altura', 'forma_vida']:
        if field in data:
            especie[field] = data[field]
    
    # Se o apelido mudou, atualizar referências na lista unificada
    old_apelido = especie.get('apelido')
    new_apelido = data.get('apelido', old_apelido)
    
    if old_apelido != new_apelido and new_apelido:
        # Atualizar na lista unificada se necessário
        if old_apelido in analysis_data['especies_unificadas']:
            # Decrementar ocorrências do apelido antigo
            analysis_data['especies_unificadas'][old_apelido]['ocorrencias'] -= 1
            if analysis_data['especies_unificadas'][old_apelido]['ocorrencias'] <= 0:
                del analysis_data['especies_unificadas'][old_apelido]
        
        # Adicionar/atualizar novo apelido
        if new_apelido not in analysis_data['especies_unificadas']:
            analysis_data['especies_unificadas'][new_apelido] = {
                'apelido_original': new_apelido,
                'apelido_usuario': new_apelido,
                'genero': data.get('genero', ''),
                'especie': data.get('especie', ''),
                'familia': data.get('familia', ''),
                'ocorrencias': 1
            }
        else:
            analysis_data['especies_unificadas'][new_apelido]['ocorrencias'] += 1
    elif new_apelido in analysis_data['especies_unificadas']:
        # Atualizar informações taxonômicas na lista unificada
        if 'genero' in data:
            analysis_data['especies_unificadas'][new_apelido]['genero'] = data['genero']
        if 'especie' in data:
            analysis_data['especies_unificadas'][new_apelido]['especie'] = data['especie']
        if 'familia' in data:
            analysis_data['especies_unificadas'][new_apelido]['familia'] = data['familia']
    
    print(f"✓ Espécie atualizada: subparcela {subparcela}, índice {index}")

    return jsonify({
        'success': True,
        'especie': especie
    })

@app.route('/api/parcela/<parcela>/subparcela/<int:subparcela>/especie/<int:index>', methods=['DELETE'])
def delete_especie_by_index(parcela, subparcela, index):
    """Remove uma espécie específica pelo índice na subparcela"""
    if parcela not in analysis_data['parcelas']:
        return jsonify({'error': 'Parcela não encontrada'}), 404

    parcela_data = analysis_data['parcelas'][parcela]
    if subparcela not in parcela_data.get('subparcelas', {}):
        return jsonify({'error': 'Subparcela não encontrada'}), 404

    subparcela_data = parcela_data['subparcelas'][subparcela]
    
    if index < 0 or index >= len(subparcela_data['especies']):
        return jsonify({'error': 'Índice de espécie inválido'}), 404

    removed = subparcela_data['especies'].pop(index)
    
    # Atualizar contagem de ocorrências
    apelido = removed['apelido']
    if apelido in analysis_data['especies_unificadas']:
        analysis_data['especies_unificadas'][apelido]['ocorrencias'] -= 1
        if analysis_data['especies_unificadas'][apelido]['ocorrencias'] <= 0:
            del analysis_data['especies_unificadas'][apelido]

    return jsonify({
        'success': True,
        'message': f'Espécie "{apelido}" removida'
    })

@app.route('/api/parcela/<parcela>/subparcela/<int:subparcela>/especie', methods=['POST'])
def add_especie_to_subparcela(parcela, subparcela):
    """Adiciona uma nova espécie à subparcela"""
    data = request.json

    if parcela not in analysis_data['parcelas']:
        return jsonify({'error': 'Parcela não encontrada'}), 404

    parcela_data = analysis_data['parcelas'][parcela]
    if subparcela not in parcela_data.get('subparcelas', {}):
        return jsonify({'error': 'Subparcela não encontrada'}), 404

    subparcela_data = parcela_data['subparcelas'][subparcela]

    nova_especie = {
        'apelido': data.get('apelido', 'Nova Espécie'),
        'genero': data.get('genero', ''),
        'familia': data.get('familia', ''),
        'observacoes': data.get('observacoes', ''),
        'cobertura': data.get('cobertura', 0),
        'altura': data.get('altura', 0),
        'forma_vida': data.get('forma_vida', 'Erva'),
        'indice': len(subparcela_data['especies']) + 1
    }
    
    subparcela_data['especies'].append(nova_especie)

    # Atualizar espécies unificadas (aninhado por parcela)
    apelido = nova_especie['apelido']

    # Garantir que especies_unificadas está aninhado por parcela
    if parcela not in analysis_data['especies_unificadas']:
        analysis_data['especies_unificadas'][parcela] = {}

    if apelido not in analysis_data['especies_unificadas'][parcela]:
        analysis_data['especies_unificadas'][parcela][apelido] = {
            'apelido_original': apelido,
            'apelido_usuario': apelido,
            'genero': data.get('genero', ''),
            'especie': '',
            'familia': data.get('familia', ''),
            'observacoes': data.get('observacoes', ''),
            'ocorrencias': 0
        }

    analysis_data['especies_unificadas'][parcela][apelido]['ocorrencias'] += 1

    print(f"✅ Espécie '{apelido}' adicionada à subparcela {subparcela} da parcela {parcela}")
    print(f"   Total de ocorrências: {analysis_data['especies_unificadas'][parcela][apelido]['ocorrencias']}")

    return jsonify({
        'success': True,
        'especie': nova_especie,
        'message': f'Espécie {apelido} adicionada com sucesso'
    })

@app.route('/api/parcela/<parcela>/especies', methods=['GET'])
def get_especies_unificadas(parcela):
    """Retorna as espécies unificadas da parcela"""
    if parcela not in analysis_data['parcelas']:
        return jsonify({'error': 'Parcela não encontrada'}), 404

    return jsonify({
        'success': True,
        'especies': analysis_data['especies_unificadas']
    })

@app.route('/api/parcela/<parcela>/subparcela/<int:subparcela>/reanalyze', methods=['POST'])
def reanalyze_subparcela(parcela, subparcela):
    """Reanalisar uma subparcela específica"""
    print(f"\n=== Reanalisando subparcela {subparcela} da parcela {parcela} ===")
    
    if parcela not in analysis_data['parcelas']:
        return jsonify({'error': 'Parcela não encontrada'}), 404
    
    parcela_info = analysis_data['parcelas'][parcela]
    
    # Encontrar a imagem da subparcela
    img_info = None
    for img in parcela_info['images']:
        if img['subparcela'] == subparcela:
            img_info = img
            break
    
    if not img_info:
        return jsonify({'error': 'Subparcela não encontrada'}), 404
    
    # Obter configuração da análise
    data = request.get_json() or {}
    ai_model = data.get('ai_model', app.config['DEFAULT_AI'])
    template_config = data.get('template_config', {'template': 'default', 'params': None})
    
    print(f"Modelo: {ai_model}, Template: {template_config.get('template', 'default')}")
    
    # Obter apelidos existentes para padronização
    existing_apelidos = list(analysis_data['especies_unificadas'].keys())
    print(f"Apelidos existentes para referência: {existing_apelidos}")
    
    # Adicionar lista de apelidos existentes aos parâmetros customizados
    if template_config.get('params') is None:
        template_config['params'] = {}
    template_config['params']['existing_species'] = existing_apelidos
    
    # Obter versão do Gemini
    gemini_version = request.headers.get('X-Gemini-Version', 'gemini-flash-latest')
    
    # Obter versão do Claude
    claude_version = request.headers.get('X-Claude-Version', 'claude-sonnet-4-5-20250929')
    
    # Obter API key (decodificar de Base64)
    api_key = None
    if ai_model == 'claude':
        api_key = decode_api_key(request.headers.get('X-API-Key-Claude'))
    elif ai_model == 'gpt4':
        api_key = decode_api_key(request.headers.get('X-API-Key-GPT4'))
    elif ai_model == 'gemini':
        api_key = decode_api_key(request.headers.get('X-API-Key-Gemini'))
    elif ai_model == 'deepseek':
        api_key = decode_api_key(request.headers.get('X-API-Key-DeepSeek'))
    elif ai_model == 'qwen':
        api_key = decode_api_key(request.headers.get('X-API-Key-Qwen'))
    elif ai_model == 'huggingface':
        api_key = decode_api_key(request.headers.get('X-API-Key-HuggingFace'))
    
    if not api_key:
        return jsonify({'error': f'API key não configurada para {ai_model}'}), 400
    
    try:
        filepath = img_info['path']
        print(f"Reanalisando: {filepath}")
        
        # Analisar com retry se necessário
        max_retries = 2
        especies_validas = []
        
        for retry in range(max_retries + 1):
            if retry > 0:
                print(f"⚠️ Tentativa {retry + 1}/{max_retries + 1}")
            
            if ai_model == 'gemini':
                analysis = analyze_image_with_ai(filepath, ai_model, api_key, gemini_version, None, template_config)
            elif ai_model == 'claude':
                analysis = analyze_image_with_ai(filepath, ai_model, api_key, None, claude_version, template_config)
            else:
                analysis = analyze_image_with_ai(filepath, ai_model, api_key, None, None, template_config)
            
            # Validar espécies
            especies_validas = []
            for esp in analysis.get('especies', []):
                apelido = esp.get('apelido', '')
                
                erro_keywords = [
                    'erro', 'error', 'falha', 'limite', 'atingido', 
                    'aguarde', 'não disponível', 'timeout', 'quota',
                    'exceeded', 'rate limit', 'api key', 'invalid'
                ]
                
                apelido_lower = apelido.lower()
                is_error = any(keyword in apelido_lower for keyword in erro_keywords)
                
                if not is_error and 'erro' not in esp:
                    especies_validas.append(esp)
            
            if len(especies_validas) > 0:
                break
            
            if retry == max_retries:
                especies_validas = [{
                    'apelido': 'Vegetação Não Detectada',
                    'genero': '',
                    'familia': '',
                    'observacoes': 'IA não conseguiu identificar vegetação clara após múltiplas tentativas.',
                    'cobertura': 100,
                    'altura': 0,
                    'forma_vida': '-'
                }]
        
        # Processar espécies (remover antigas da contagem)
        old_especies = parcela_info['subparcelas'].get(subparcela, {}).get('especies', [])
        for old_esp in old_especies:
            old_apelido = old_esp['apelido']
            if old_apelido in analysis_data['especies_unificadas']:
                analysis_data['especies_unificadas'][old_apelido]['ocorrencias'] -= 1
                if analysis_data['especies_unificadas'][old_apelido]['ocorrencias'] <= 0:
                    del analysis_data['especies_unificadas'][old_apelido]
        
        # Adicionar novas espécies
        especies_encontradas = []
        for esp_idx, esp in enumerate(especies_validas, 1):
            apelido = esp['apelido']
            
            if apelido not in analysis_data['especies_unificadas']:
                analysis_data['especies_unificadas'][apelido] = {
                    'apelido_original': apelido,
                    'apelido_usuario': apelido,
                    'genero': esp.get('genero', ''),
                    'especie': '',
                    'familia': esp.get('familia', ''),
                    'observacoes': esp.get('observacoes', ''),
                    'ocorrencias': 0
                }
            else:
                if esp.get('genero') and not analysis_data['especies_unificadas'][apelido]['genero']:
                    analysis_data['especies_unificadas'][apelido]['genero'] = esp.get('genero')
                if esp.get('familia') and not analysis_data['especies_unificadas'][apelido]['familia']:
                    analysis_data['especies_unificadas'][apelido]['familia'] = esp.get('familia')
                if esp.get('observacoes') and not analysis_data['especies_unificadas'][apelido]['observacoes']:
                    analysis_data['especies_unificadas'][apelido]['observacoes'] = esp.get('observacoes')
            
            analysis_data['especies_unificadas'][apelido]['ocorrencias'] += 1
            
            especies_encontradas.append({
                'indice': esp_idx,
                'apelido': apelido,
                'genero': esp.get('genero', ''),
                'familia': esp.get('familia', ''),
                'observacoes': esp.get('observacoes', ''),
                'cobertura': esp['cobertura'],
                'altura': esp['altura'],
                'forma_vida': esp['forma_vida']
            })
        
        # Atualizar subparcela
        parcela_info['subparcelas'][subparcela] = {
            'image': img_info['filename'],
            'especies': especies_encontradas
        }
        
        print(f"✓ Reanálise concluída: {len(especies_encontradas)} espécies")
        
        # Retornar apenas espécies da parcela atual (aninhadas por parcela se existir)
        especies_para_retornar = {}
        if parcela in analysis_data.get('especies_unificadas', {}):
            # Estrutura aninhada por parcela
            especies_para_retornar = analysis_data['especies_unificadas'][parcela]
        else:
            # Estrutura global - filtrar apenas espécies desta parcela
            # Verificar quais espécies aparecem nesta parcela
            especies_desta_parcela = set()
            for sub_id, sub_data in parcela_info['subparcelas'].items():
                for esp in sub_data.get('especies', []):
                    especies_desta_parcela.add(esp['apelido'])
            
            # Retornar apenas essas espécies
            especies_para_retornar = {
                apelido: info 
                for apelido, info in analysis_data.get('especies_unificadas', {}).items()
                if apelido in especies_desta_parcela
            }
        
        return jsonify({
            'success': True,
            'subparcela': subparcela,
            'especies': especies_encontradas,
            'especies_unificadas': especies_para_retornar
        })
        
    except Exception as e:
        error_msg = str(e)
        print(f"ERRO na reanálise: {error_msg}")
        return jsonify({'error': error_msg}), 500

@app.route('/api/especies/<parcela>/<int:subparcela>/coverage', methods=['PUT'])
def save_coverage_data(parcela, subparcela):
    """Salva dados de cobertura visual (formas desenhadas) e atualiza porcentagens"""
    data = request.json

    if parcela not in analysis_data['parcelas']:
        return jsonify({'error': 'Parcela não encontrada'}), 404

    parcela_data = analysis_data['parcelas'][parcela]
    if subparcela not in parcela_data.get('subparcelas', {}):
        return jsonify({'error': 'Subparcela não encontrada'}), 404

    subparcela_data = parcela_data['subparcelas'][subparcela]

    try:
        # Salvar dados de cobertura (formas geométricas)
        coverage_data = data.get('coverageData', {})
        subparcela_data['coverageData'] = coverage_data

        # Atualizar porcentagens de cobertura das espécies
        especies_updates = data.get('especies', [])
        for update in especies_updates:
            apelido = update.get('apelido')
            nova_cobertura = update.get('cobertura')

            if apelido and nova_cobertura is not None:
                # Encontrar e atualizar a espécie
                for esp in subparcela_data['especies']:
                    if esp['apelido'] == apelido:
                        esp['cobertura'] = nova_cobertura
                        break

        print(f"✓ Dados de cobertura salvos para subparcela {subparcela}")
        print(f"  - Área da subparcela: {'definida' if coverage_data.get('subparcelaShape') else 'não definida'}")
        print(f"  - Áreas de espécies: {len(coverage_data.get('speciesShapes', []))} espécie(s) com áreas definidas")

        return jsonify({
            'success': True,
            'message': 'Dados de cobertura salvos com sucesso'
        })

    except Exception as e:
        error_msg = str(e)
        print(f"ERRO ao salvar dados de cobertura: {error_msg}")
        return jsonify({'error': error_msg}), 500

@app.route('/api/species/coverage', methods=['POST'])
def update_species_coverage():
    """Atualiza porcentagem de cobertura de uma espécie específica"""
    data = request.json
    
    subparcela_id = data.get('subparcela_id')
    especie_nome = data.get('especie_nome')
    apelido = data.get('apelido')
    cobertura = data.get('cobertura')
    
    if not subparcela_id or cobertura is None:
        return jsonify({'error': 'Dados insuficientes'}), 400
    
    try:
        # Buscar em todas as parcelas
        for parcela_nome, parcela_data in analysis_data['parcelas'].items():
            if subparcela_id in parcela_data.get('subparcelas', {}):
                subparcela_data = parcela_data['subparcelas'][subparcela_id]
                
                # Atualizar cobertura da espécie
                for esp in subparcela_data.get('especies', []):
                    if (esp.get('especie') == especie_nome or 
                        esp.get('apelido') == apelido):
                        esp['cobertura'] = float(cobertura)
                        print(f"✓ Cobertura atualizada: {apelido} = {cobertura}%")
                        return jsonify({'success': True})
                
        return jsonify({'error': 'Espécie não encontrada'}), 404
        
    except Exception as e:
        print(f"ERRO ao atualizar cobertura: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/species/area', methods=['POST'])
def update_species_area():
    """Atualiza áreas/polígonos desenhados de uma espécie"""
    data = request.json
    
    parcela_nome = data.get('parcela')
    subparcela_id = data.get('subparcela')
    especie_nome = data.get('especie')
    area_shapes = data.get('area_shapes', [])
    
    if not parcela_nome or not subparcela_id or not especie_nome:
        return jsonify({'error': 'Dados insuficientes'}), 400
    
    try:
        if parcela_nome not in analysis_data['parcelas']:
            return jsonify({'error': 'Parcela não encontrada'}), 404
            
        parcela_data = analysis_data['parcelas'][parcela_nome]
        
        if subparcela_id not in parcela_data.get('subparcelas', {}):
            return jsonify({'error': 'Subparcela não encontrada'}), 404
        
        subparcela_data = parcela_data['subparcelas'][subparcela_id]
        
        # Atualizar área da espécie
        for esp in subparcela_data.get('especies', []):
            if esp.get('apelido') == especie_nome or esp.get('especie') == especie_nome:
                esp['area_shapes'] = area_shapes
                print(f"✓ Áreas da espécie {especie_nome} atualizadas: {len(area_shapes)} polígonos")
                return jsonify({'success': True})
        
        return jsonify({'error': 'Espécie não encontrada'}), 404
        
    except Exception as e:
        print(f"ERRO ao atualizar área: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/parcela/<parcela>/subparcela/<int:subparcela>/add-species-ai', methods=['POST'])
def add_species_with_ai(parcela, subparcela):
    """Analisa imagem com IA para detectar espécies adicionais"""
    print(f"\n=== Analisando espécies adicionais com IA - {parcela}/{subparcela} ===")

    if parcela not in analysis_data['parcelas']:
        return jsonify({'error': 'Parcela não encontrada'}), 404

    parcela_info = analysis_data['parcelas'][parcela]

    # Encontrar a imagem da subparcela
    img_info = None
    for img in parcela_info['images']:
        if img['subparcela'] == subparcela:
            img_info = img
            break

    if not img_info:
        return jsonify({'error': 'Subparcela não encontrada'}), 404

    # Obter configuração
    data = request.get_json() or {}
    ai_model = data.get('ai_model', app.config['DEFAULT_AI'])
    existing_species = data.get('existing_species', [])

    print(f"Modelo: {ai_model}")
    print(f"Espécies existentes: {existing_species}")

    # Obter API keys
    api_key = None
    gemini_version = request.headers.get('X-Gemini-Version', 'gemini-flash-latest')
    claude_version = request.headers.get('X-Claude-Version', 'claude-sonnet-4-5-20250929')

    if ai_model == 'claude':
        api_key = decode_api_key(request.headers.get('X-API-Key-Claude'))
    elif ai_model == 'gpt4':
        api_key = decode_api_key(request.headers.get('X-API-Key-GPT4'))
    elif ai_model == 'gemini':
        api_key = decode_api_key(request.headers.get('X-API-Key-Gemini'))
    elif ai_model == 'deepseek':
        api_key = decode_api_key(request.headers.get('X-API-Key-DeepSeek'))
    elif ai_model == 'qwen':
        api_key = decode_api_key(request.headers.get('X-API-Key-Qwen'))
    elif ai_model == 'huggingface':
        api_key = decode_api_key(request.headers.get('X-API-Key-HuggingFace'))

    if not api_key:
        return jsonify({'error': f'API key não configurada para {ai_model}'}), 400

    try:
        filepath = img_info['path']
        print(f"Analisando: {filepath}")

        # Criar template customizado que instrui a IA a procurar espécies adicionais
        template_config = {
            'template': 'default',
            'params': {
                'existing_species': existing_species,
                'detect_coordinates': True,
                'min_species': 1,
                'max_species': 5  # Limitar a 5 novas espécies
            }
        }

        # Analisar
        if ai_model == 'gemini':
            analysis = analyze_image_with_ai(filepath, ai_model, api_key, gemini_version, None, template_config)
        elif ai_model == 'claude':
            analysis = analyze_image_with_ai(filepath, ai_model, api_key, None, claude_version, template_config)
        else:
            analysis = analyze_image_with_ai(filepath, ai_model, api_key, None, None, template_config)

        # Filtrar apenas espécies novas (não detectadas anteriormente)
        new_species = []
        for esp in analysis.get('especies', []):
            apelido = esp.get('apelido', '')

            # Ignorar se já existe
            if apelido in existing_species:
                print(f"   ⚠️ Espécie '{apelido}' já existe, ignorando")
                continue

            # Ignorar erros
            erro_keywords = [
                'erro', 'error', 'falha', 'limite', 'atingido',
                'aguarde', 'não disponível', 'timeout', 'quota',
                'exceeded', 'rate limit', 'api key', 'invalid'
            ]

            if any(keyword in apelido.lower() for keyword in erro_keywords):
                print(f"   ❌ Ignorando erro: {apelido}")
                continue

            new_species.append(esp)
            print(f"   ✅ Nova espécie detectada: {apelido}")

        if len(new_species) == 0:
            return jsonify({
                'success': True,
                'new_species': [],
                'message': 'Nenhuma espécie adicional detectada'
            })

        # Adicionar novas espécies à subparcela
        subparcela_data = parcela_info['subparcelas'][subparcela]

        for esp in new_species:
            apelido = esp['apelido']

            # Adicionar à subparcela
            nova_especie = {
                'indice': len(subparcela_data['especies']) + 1,
                'apelido': apelido,
                'genero': esp.get('genero', ''),
                'familia': esp.get('familia', ''),
                'observacoes': esp.get('observacoes', ''),
                'cobertura': esp.get('cobertura', 5),
                'altura': esp.get('altura', 10),
                'forma_vida': esp.get('forma_vida', 'Erva'),
                'areas': esp.get('areas', [])  # Coordenadas da IA
            }

            subparcela_data['especies'].append(nova_especie)

            # Atualizar espécies unificadas
            if apelido not in analysis_data['especies_unificadas']:
                analysis_data['especies_unificadas'][apelido] = {
                    'apelido_original': apelido,
                    'apelido_usuario': apelido,
                    'genero': esp.get('genero', ''),
                    'especie': '',
                    'familia': esp.get('familia', ''),
                    'observacoes': esp.get('observacoes', ''),
                    'ocorrencias': 0
                }

            analysis_data['especies_unificadas'][apelido]['ocorrencias'] += 1

        print(f"✓ {len(new_species)} novas espécies adicionadas")

        return jsonify({
            'success': True,
            'new_species': new_species,
            'message': f'{len(new_species)} nova(s) espécie(s) detectada(s)'
        })

    except Exception as e:
        error_msg = str(e)
        print(f"ERRO na análise: {error_msg}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': error_msg}), 500

@app.route('/api/export', methods=['POST'])
def export_excel():
    """Exporta dados completos para Excel com análises profissionais"""
    try:
        data = request.json
        parcela_nome = data.get('parcela', 'Parcela_1')
        
        print(f"\n📊 Iniciando exportação Excel para parcela: {parcela_nome}")
        print(f"📦 Dados recebidos: {json.dumps(data, indent=2, ensure_ascii=False)}")

        if parcela_nome not in analysis_data['parcelas']:
            return jsonify({'error': 'Parcela não encontrada'}), 404

        parcela = analysis_data['parcelas'][parcela_nome]
        subparcelas_data = data.get('subparcelas', [])
        especies_unificadas = data.get('especies_unificadas', [])
        estatisticas = data.get('estatisticas', {})

        # Criar workbook
        wb = openpyxl.Workbook()
        
        # ==== ABA 1: RESUMO EXECUTIVO ====
        ws_resumo = wb.active
        ws_resumo.title = "Resumo Executivo"
        
        # Título principal
        ws_resumo['A1'] = f"RELATÓRIO DE ANÁLISE - {parcela_nome}"
        ws_resumo['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_resumo['A1'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        ws_resumo['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_resumo.merge_cells('A1:E1')
        ws_resumo.row_dimensions[1].height = 30
        
        # Informações gerais
        ws_resumo['A3'] = "Data da Análise:"
        ws_resumo['B3'] = data.get('data_analise', datetime.now().strftime("%d/%m/%Y"))
        ws_resumo['A4'] = "Total de Subparcelas:"
        ws_resumo['B4'] = estatisticas.get('total_subparcelas', len(subparcelas_data))
        ws_resumo['A5'] = "Espécies Únicas:"
        ws_resumo['B5'] = estatisticas.get('total_especies_unicas', len(especies_unificadas))
        ws_resumo['A6'] = "Cobertura Total (%):"
        ws_resumo['B6'] = round(estatisticas.get('cobertura_total', 0), 2)
        ws_resumo['A7'] = "Altura Média (cm):"
        ws_resumo['B7'] = round(estatisticas.get('altura_media', 0), 2)
        
        for row in range(3, 8):
            ws_resumo[f'A{row}'].font = Font(bold=True)
            ws_resumo[f'A{row}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        
        # ==== ABA 2: DADOS DETALHADOS ====
        ws_detalhes = wb.create_sheet("Dados Detalhados")
        
        headers = ['Subparcela', 'Índice', 'Apelido', 'Gênero', 'Família', 
                   'Cobertura (%)', 'Altura (cm)', 'Forma de Vida', 'Observações']
        ws_detalhes.append(headers)
        
        # Formatação do cabeçalho
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws_detalhes[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adicionar dados de cada subparcela
        for subparcela in subparcelas_data:
            sub_num = subparcela.get('numero')
            for idx, esp in enumerate(subparcela.get('especies', []), 1):
                row_data = [
                    sub_num,
                    idx,
                    esp.get('apelido', ''),
                    esp.get('genero', ''),
                    esp.get('familia', ''),
                    round(esp.get('cobertura', 0), 2),
                    round(esp.get('altura', 0), 2),
                    esp.get('forma_vida', ''),
                    esp.get('observacoes', '')
                ]
                ws_detalhes.append(row_data)
        
        # Ajustar larguras
        column_widths = [12, 8, 25, 15, 20, 12, 12, 14, 30]
        for idx, width in enumerate(column_widths, 1):
            ws_detalhes.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
        
        # ==== ABA 3: ESPÉCIES UNIFICADAS ====
        ws_especies = wb.create_sheet("Espécies Unificadas")
        
        especies_headers = ['Apelido Original', 'Apelido Usuário', 'Gênero', 'Espécie',
                            'Família', 'Nº Ocorrências', 'Observações']
        ws_especies.append(especies_headers)
        
        for cell in ws_especies[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for esp in especies_unificadas:
            ws_especies.append([
                esp.get('apelido_original', ''),
                esp.get('apelido_usuario', ''),
                esp.get('genero', ''),
                esp.get('especie', ''),
                esp.get('familia', ''),
                esp.get('ocorrencias', 0),
                esp.get('observacoes', '')
            ])
        
        especies_widths = [25, 25, 15, 20, 20, 14, 30]
        for idx, width in enumerate(especies_widths, 1):
            ws_especies.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
        
        # ==== ABA 4: ANÁLISE POR SUBPARCELA ====
        ws_analise = wb.create_sheet("Análise por Subparcela")
        
        analise_headers = ['Subparcela', 'Total Espécies', 'Cobertura Total (%)', 
                           'Altura Média (cm)', 'Forma de Vida Dominante']
        ws_analise.append(analise_headers)
        
        for cell in ws_analise[1]:
            cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for subparcela in subparcelas_data:
            especies = subparcela.get('especies', [])
            total_esp = len(especies)
            cobertura_total = sum(e.get('cobertura', 0) for e in especies)
            altura_media = sum(e.get('altura', 0) for e in especies) / total_esp if total_esp > 0 else 0
            
            # Forma de vida dominante
            formas_vida = [e.get('forma_vida', 'Erva') for e in especies]
            forma_dominante = max(set(formas_vida), key=formas_vida.count) if formas_vida else 'N/A'
            
            ws_analise.append([
                subparcela.get('numero'),
                total_esp,
                round(cobertura_total, 2),
                round(altura_media, 2),
                forma_dominante
            ])
        
        analise_widths = [12, 14, 18, 16, 22]
        for idx, width in enumerate(analise_widths, 1):
            ws_analise.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
        
        # ==== ABA 5: DIVERSIDADE E ESTATÍSTICAS ====
        ws_stats = wb.create_sheet("Estatísticas")
        
        ws_stats['A1'] = "ESTATÍSTICAS DE DIVERSIDADE"
        ws_stats['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_stats['A1'].fill = PatternFill(start_color="673AB7", end_color="673AB7", fill_type="solid")
        ws_stats.merge_cells('A1:B1')
        
        # Estatísticas gerais
        stats_data = [
            ["Métrica", "Valor"],
            ["Riqueza de Espécies", len(especies_unificadas)],
            ["Total de Registros", sum(e.get('ocorrencias', 0) for e in especies_unificadas)],
            ["Cobertura Média por Espécie (%)", round(estatisticas.get('cobertura_total', 0) / len(especies_unificadas), 2) if especies_unificadas else 0],
            ["Altura Média Geral (cm)", round(estatisticas.get('altura_media', 0), 2)],
            ["Espécie Mais Frequente", max(especies_unificadas, key=lambda e: e.get('ocorrencias', 0)).get('apelido_original', 'N/A') if especies_unificadas else 'N/A'],
        ]
        
        for row_idx, row_data in enumerate(stats_data, 3):
            ws_stats.append(row_data)
            if row_idx == 3:  # Header
                ws_stats[f'A{row_idx}'].font = Font(bold=True)
                ws_stats[f'B{row_idx}'].font = Font(bold=True)
                ws_stats[f'A{row_idx}'].fill = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")
                ws_stats[f'B{row_idx}'].fill = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")
        
        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 20
        
        # Salvar arquivo
        os.makedirs('exports', exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{parcela_nome}_Relatorio_Completo_{timestamp}.xlsx"
        filepath = os.path.join('exports', filename)
        wb.save(filepath)
        
        print(f"✅ Arquivo Excel salvo: {filepath}")

        return jsonify({
            'success': True,
            'filename': filename,
            'download_url': f'/api/download/{filename}'
        })
        
    except Exception as e:
        print(f"❌ Erro na exportação: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/download/<filename>')
def download_file(filename):
    """Download do arquivo Excel exportado"""
    filepath = os.path.join('exports', filename)
    if not os.path.exists(filepath):
        return jsonify({'error': 'Arquivo não encontrado'}), 404

    return send_file(filepath, as_attachment=True, download_name=filename)

@app.route('/api/parcelas', methods=['GET'])
def get_parcelas():
    """Retorna lista de parcelas"""
    parcelas_list = []
    for nome, dados in analysis_data['parcelas'].items():
        parcelas_list.append({
            'nome': nome,
            'num_subparcelas': len(dados.get('subparcelas', {})),
            'num_imagens': len(dados.get('images', []))
        })

    return jsonify({'parcelas': parcelas_list})

@app.route('/api/parcela/<nome>', methods=['GET'])
def get_parcela_details(nome):
    """Retorna detalhes de uma parcela específica"""
    if nome not in analysis_data['parcelas']:
        return jsonify({'error': 'Parcela não encontrada'}), 404

    parcela = analysis_data['parcelas'][nome]
    return jsonify({
        'nome': nome,
        'subparcelas': parcela.get('subparcelas', {})
    })

@app.route('/api/config/apikey', methods=['POST'])
def save_apikey_config():
    """Endpoint para salvar configuração de API key (apenas confirmação, keys são gerenciadas no frontend)"""
    data = request.json
    return jsonify({
        'success': True,
        'message': 'API key configurada localmente'
    })

@app.route('/api/parcela/<parcela_nome>/images', methods=['GET'])
def get_parcela_images(parcela_nome):
    """Retorna lista de imagens de uma parcela para modo manual"""
    if parcela_nome not in analysis_data['parcelas']:
        return jsonify({'error': 'Parcela não encontrada'}), 404

    parcela = analysis_data['parcelas'][parcela_nome]
    images_list = []

    # CRITICAL FIX: Criar estrutura de subparcelas para modo manual
    if 'subparcelas' not in parcela:
        parcela['subparcelas'] = {}

    # Processar lista de imagens
    for idx, img_info in enumerate(parcela.get('images', []), 1):
        if isinstance(img_info, dict):
            img_path = img_info.get('path', '')
        else:
            img_path = str(img_info)

        if img_path and os.path.exists(img_path):
            images_list.append({
                'subparcela': idx,
                'path': img_path,
                'filename': os.path.basename(img_path)
            })

            # CRITICAL FIX: Criar subparcela vazia no backend se não existir
            if idx not in parcela['subparcelas']:
                parcela['subparcelas'][idx] = {
                    'nome': f'Subparcela {idx}',
                    'image': os.path.basename(img_path),
                    'especies': [],
                    'manual_mode': True
                }
                print(f"✅ Subparcela {idx} criada no backend para modo manual")

    # Garantir que especies_unificadas existe para esta parcela
    if parcela_nome not in analysis_data['especies_unificadas']:
        analysis_data['especies_unificadas'][parcela_nome] = {}

    print(f"📊 Modo manual: {len(images_list)} subparcelas preparadas para parcela {parcela_nome}")

    return jsonify({
        'success': True,
        'images': images_list,
        'total': len(images_list)
    })

@app.route('/api/clear-analysis', methods=['POST'])
def clear_analysis():
    """Limpa todos os dados de análise do backend para iniciar uma nova"""
    global analysis_data
    
    try:
        # Resetar estrutura de dados global
        analysis_data = {
            'parcelas': {},
            'especies_unificadas': {}
        }
        
        print("✓ Dados de análise limpos no backend")
        
        return jsonify({
            'success': True,
            'message': 'Análise limpa com sucesso'
        })
    except Exception as e:
        print(f"Erro ao limpar análise: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/analysis/save', methods=['POST'])
def save_analysis():
    """Salva uma análise em progresso para continuar depois"""
    try:
        data = request.get_json()
        analysis_name = data.get('name')
        parcela_name = data.get('parcela')
        
        if not analysis_name or not parcela_name:
            return jsonify({'error': 'Nome da análise e parcela são obrigatórios'}), 400
        
        if parcela_name not in analysis_data['parcelas']:
            return jsonify({'error': 'Parcela não encontrada'}), 404
        
        # Criar diretório para análises salvas
        saved_analyses_dir = os.path.join(os.path.dirname(__file__), 'saved_analyses')
        os.makedirs(saved_analyses_dir, exist_ok=True)
        
        # Sanitizar nome do arquivo
        safe_name = "".join(c for c in analysis_name if c.isalnum() or c in (' ', '-', '_')).strip()
        analysis_file = os.path.join(saved_analyses_dir, f"{safe_name}.json")
        
        # Dados da parcela
        parcela_data = analysis_data['parcelas'][parcela_name]
        
        # Copiar imagens para dentro da análise salva (opcional - pode ser grande)
        # Por ora, vamos salvar apenas os caminhos relativos
        images_info = []
        for img in parcela_data.get('images', []):
            # img pode ser string (path) ou dict com 'path'
            if isinstance(img, dict):
                img_path = img.get('path', '')
            else:
                img_path = str(img)
            
            if img_path and os.path.exists(img_path):
                images_info.append({
                    'path': img_path,
                    'filename': os.path.basename(img_path)
                })
        
        # Preparar dados completos da análise
        analysis_save_data = {
            'name': analysis_name,
            'parcela': parcela_name,
            'saved_at': datetime.now().isoformat(),
            'version': '2.0',
            'data': {
                'subparcelas': parcela_data.get('subparcelas', {}),
                'especies_unificadas': analysis_data.get('especies_unificadas', {}),
                'images': images_info,
                'config': data.get('config', {})  # Configurações usadas
            }
        }
        
        # Salvar como JSON
        with open(analysis_file, 'w', encoding='utf-8') as f:
            json.dump(analysis_save_data, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            'success': True,
            'message': f'Análise "{analysis_name}" salva com sucesso',
            'filename': f"{safe_name}.json"
        })
    except Exception as e:
        print(f"Erro ao salvar análise: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/analysis/list', methods=['GET'])
def list_saved_analyses():
    """Lista todas as análises salvas"""
    try:
        saved_analyses_dir = os.path.join(os.path.dirname(__file__), 'saved_analyses')
        
        if not os.path.exists(saved_analyses_dir):
            return jsonify({'analyses': []})
        
        analyses = []
        for filename in os.listdir(saved_analyses_dir):
            if filename.endswith('.json'):
                filepath = os.path.join(saved_analyses_dir, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        analysis_data_file = json.load(f)
                        
                        # Contar espécies
                        num_especies = len(analysis_data_file.get('data', {}).get('especies_unificadas', {}))
                        num_subparcelas = len(analysis_data_file.get('data', {}).get('subparcelas', {}))
                        
                        analyses.append({
                            'filename': filename,
                            'name': analysis_data_file.get('name', filename),
                            'parcela': analysis_data_file.get('parcela', 'N/A'),
                            'saved_at': analysis_data_file.get('saved_at'),
                            'num_especies': num_especies,
                            'num_subparcelas': num_subparcelas,
                            'num_images': len(analysis_data_file.get('data', {}).get('images', []))
                        })
                except Exception as e:
                    print(f"Erro ao ler análise {filename}: {e}")
                    continue
        
        # Ordenar por data de salvamento (mais recente primeiro)
        analyses.sort(key=lambda x: x.get('saved_at', ''), reverse=True)
        
        return jsonify({'analyses': analyses})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/analysis/load/<filename>', methods=['GET'])
def load_saved_analysis(filename):
    """Carrega uma análise salva"""
    try:
        saved_analyses_dir = os.path.join(os.path.dirname(__file__), 'saved_analyses')
        filepath = os.path.join(saved_analyses_dir, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'Análise não encontrada'}), 404
        
        with open(filepath, 'r', encoding='utf-8') as f:
            analysis_save_data = json.load(f)
        
        # Restaurar dados na memória
        parcela_name = analysis_save_data.get('parcela')
        
        # Verificar se as imagens ainda existem
        missing_images = []
        for img_info in analysis_save_data.get('data', {}).get('images', []):
            if not os.path.exists(img_info['path']):
                missing_images.append(img_info['filename'])
        
        return jsonify({
            'success': True,
            'analysis': analysis_save_data,
            'missing_images': missing_images
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/analysis/restore', methods=['POST'])
def restore_analysis():
    """Restaura uma análise salva para o estado atual de trabalho"""
    try:
        data = request.get_json()
        filename = data.get('filename')
        
        if not filename:
            return jsonify({'error': 'Filename é obrigatório'}), 400
        
        saved_analyses_dir = os.path.join(os.path.dirname(__file__), 'saved_analyses')
        filepath = os.path.join(saved_analyses_dir, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'Análise não encontrada'}), 404
        
        with open(filepath, 'r', encoding='utf-8') as f:
            analysis_save_data = json.load(f)
        
        parcela_name = analysis_save_data.get('parcela')
        saved_data = analysis_save_data.get('data', {})
        
        # Restaurar parcela na memória global
        if parcela_name not in analysis_data['parcelas']:
            analysis_data['parcelas'][parcela_name] = {
                'images': [],
                'subparcelas': {}
            }
        
        # Restaurar imagens com estrutura correta
        images_list = []
        for img_info in saved_data.get('images', []):
            if isinstance(img_info, dict):
                if os.path.exists(img_info.get('path', '')):
                    images_list.append(img_info)
            elif isinstance(img_info, str) and os.path.exists(img_info):
                # Reconstruir estrutura se for apenas string
                images_list.append({
                    'filename': os.path.basename(img_info),
                    'path': img_info,
                    'subparcela': len(images_list) + 1
                })
        
        analysis_data['parcelas'][parcela_name]['images'] = images_list
        
        # Restaurar subparcelas e espécies
        analysis_data['parcelas'][parcela_name]['subparcelas'] = saved_data.get('subparcelas', {})
        
        # Restaurar espécies unificadas (corrigir aninhamento duplo)
        especies_data = saved_data.get('especies_unificadas', {})
        if parcela_name in especies_data:
            # Já está aninhado por parcela no arquivo salvo
            analysis_data['especies_unificadas'][parcela_name] = especies_data[parcela_name]
        else:
            # Dados diretos (formato antigo)
            analysis_data['especies_unificadas'][parcela_name] = especies_data
        
        print(f"✓ Análise restaurada: {parcela_name}")
        print(f"✓ Images: {len(images_list)}")
        print(f"✓ Subparcelas: {len(saved_data.get('subparcelas', {}))}")
        print(f"✓ Espécies: {len(analysis_data['especies_unificadas'][parcela_name])}")
        
        return jsonify({
            'success': True,
            'message': f'Análise restaurada com sucesso',
            'parcela': parcela_name
        })
    except Exception as e:
        print(f"Erro ao restaurar análise: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/analysis/delete/<filename>', methods=['DELETE'])
def delete_saved_analysis(filename):
    """Deleta uma análise salva"""
    try:
        saved_analyses_dir = os.path.join(os.path.dirname(__file__), 'saved_analyses')
        filepath = os.path.join(saved_analyses_dir, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'Análise não encontrada'}), 404
        
        os.remove(filepath)
        
        return jsonify({
            'success': True,
            'message': 'Análise deletada com sucesso'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reference-species', methods=['GET'])
def get_reference_species():
    """Retorna a lista de espécies de referência"""
    try:
        ref_file = os.path.join(os.path.dirname(__file__), 'reference_species.json')
        
        if os.path.exists(ref_file):
            with open(ref_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return jsonify({'species': data.get('species', [])})
        else:
            return jsonify({'species': []})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reference-species', methods=['POST'])
def save_reference_species():
    """Salva a lista de espécies de referência"""
    try:
        data = request.get_json()
        species_list = data.get('species', [])
        
        ref_file = os.path.join(os.path.dirname(__file__), 'reference_species.json')
        
        with open(ref_file, 'w', encoding='utf-8') as f:
            json.dump({
                'species': species_list,
                'updated_at': datetime.now().isoformat()
            }, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            'success': True,
            'message': f'{len(species_list)} espécies salvas com sucesso'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reference-species/<int:index>', methods=['DELETE'])
def delete_reference_species(index):
    """Deleta uma espécie de referência pelo índice"""
    try:
        ref_file = os.path.join(os.path.dirname(__file__), 'reference_species.json')
        
        if not os.path.exists(ref_file):
            return jsonify({'error': 'Arquivo não encontrado'}), 404
        
        with open(ref_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        species_list = data.get('species', [])
        
        if index < 0 or index >= len(species_list):
            return jsonify({'error': 'Índice inválido'}), 400
        
        removed = species_list.pop(index)
        
        with open(ref_file, 'w', encoding='utf-8') as f:
            json.dump({
                'species': species_list,
                'updated_at': datetime.now().isoformat()
            }, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            'success': True,
            'message': f'Espécie "{removed.get("apelido", "")}" removida'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/analysis/export-complete', methods=['POST'])
def export_complete_analysis():
    """Exporta análise completa em arquivo ZIP com JSON e todas as imagens"""
    try:
        data = request.get_json()
        parcela_name = data.get('parcela')
        
        if not parcela_name or parcela_name not in analysis_data['parcelas']:
            return jsonify({'error': 'Parcela não encontrada'}), 404
        
        parcela_data = analysis_data['parcelas'][parcela_name]
        
        # Criar ZIP em memória
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # 1. Salvar JSON com dados completos
            analysis_json = {
                'version': '2.0',
                'exported_at': datetime.now().isoformat(),
                'parcela': parcela_name,
                'subparcelas': parcela_data.get('subparcelas', {}),
                'especies_unificadas': analysis_data['especies_unificadas'].get(parcela_name, {}),
                'metadata': {
                    'num_subparcelas': len(parcela_data.get('subparcelas', {})),
                    'num_especies': len(analysis_data['especies_unificadas'].get(parcela_name, {})),
                    'num_imagens': len(parcela_data.get('images', []))
                }
            }
            
            zip_file.writestr('analysis_data.json', json.dumps(analysis_json, ensure_ascii=False, indent=2))
            
            # 2. Copiar todas as imagens referenciadas
            images_copied = 0
            image_paths = set()
            
            # Coletar todas as imagens das subparcelas
            for subparcela_id, subparcela in parcela_data.get('subparcelas', {}).items():
                img_path = subparcela.get('image_path')
                if img_path and os.path.exists(img_path):
                    image_paths.add(img_path)
            
            # Adicionar imagens da parcela (se houver)
            for img in parcela_data.get('images', []):
                if isinstance(img, dict):
                    img_path = img.get('path')
                else:
                    img_path = str(img)
                    
                if img_path and os.path.exists(img_path):
                    image_paths.add(img_path)
            
            # Copiar imagens para o ZIP
            for img_path in image_paths:
                try:
                    # Manter estrutura de pastas relativa
                    filename = os.path.basename(img_path)
                    zip_file.write(img_path, f'images/{filename}')
                    images_copied += 1
                except Exception as e:
                    print(f"Erro ao adicionar imagem {img_path}: {e}")
            
            # 3. Adicionar README com instruções
            readme = f"""# Análise Exportada: {parcela_name}

Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}

## Conteúdo

- `analysis_data.json`: Dados completos da análise
- `images/`: {images_copied} imagens das subparcelas

## Como Importar

1. Abra a plataforma de Análise de Vegetação Herbácea
2. Clique em "📂 Carregar Análise" (topo da página)
3. Clique em "📥 Importar ZIP Completo"
4. Selecione este arquivo ZIP
5. Aguarde o processamento

## Estatísticas

- Subparcelas: {analysis_json['metadata']['num_subparcelas']}
- Espécies identificadas: {analysis_json['metadata']['num_especies']}
- Imagens incluídas: {images_copied}

## Observações

Este arquivo contém TODOS os dados, imagens e anotações da análise.
Pode ser compartilhado com outros usuários da plataforma.
"""
            zip_file.writestr('README.txt', readme)
        
        # Preparar para download
        zip_buffer.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{parcela_name}_{timestamp}.zip"
        
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Erro ao exportar análise completa: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/analysis/import-complete', methods=['POST'])
def import_complete_analysis():
    """Importa análise completa de arquivo ZIP"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        zip_file = request.files['file']
        
        if not zip_file.filename.endswith('.zip'):
            return jsonify({'error': 'Arquivo deve ser um ZIP'}), 400
        
        # Criar diretório temporário
        temp_dir = os.path.join(os.path.dirname(__file__), 'temp_import')
        os.makedirs(temp_dir, exist_ok=True)
        
        try:
            # Extrair ZIP
            with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                # Listar conteúdo do ZIP para debug
                print("📦 Conteúdo do ZIP:")
                for name in zip_ref.namelist():
                    print(f"   - {name}")

                zip_ref.extractall(temp_dir)

            # Listar arquivos extraídos para debug
            print(f"\n📁 Arquivos em {temp_dir}:")
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    rel_path = os.path.relpath(os.path.join(root, file), temp_dir)
                    print(f"   - {rel_path}")

            # Procurar arquivo JSON - pode ser {parcela}_dados.json ou analysis_data.json
            json_path = None
            json_patterns = ['*_dados.json', 'analysis_data.json']
            
            for pattern in json_patterns:
                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        if file.endswith('_dados.json') or file == 'analysis_data.json':
                            json_path = os.path.join(root, file)
                            print(f"✓ JSON encontrado: {json_path}")
                            break
                    if json_path:
                        break
                if json_path:
                    break
            
            if not json_path:
                return jsonify({'error': 'Arquivo JSON de dados não encontrado no ZIP. Procurado: *_dados.json ou analysis_data.json'}), 400
            
            with open(json_path, 'r', encoding='utf-8') as f:
                imported_data = json.load(f)
            
            # Validar estrutura do JSON
            if 'parcela' not in imported_data:
                return jsonify({'error': 'JSON inválido: campo "parcela" não encontrado'}), 400
            
            # Aceitar tanto formato novo (subparcelas) quanto legado (analysisResults)
            if 'subparcelas' not in imported_data and 'analysisResults' not in imported_data:
                return jsonify({'error': 'JSON inválido: nem "subparcelas" nem "analysisResults" encontrado'}), 400
            
            # Converter formato legado para novo se necessário
            if 'analysisResults' in imported_data and 'subparcelas' not in imported_data:
                print("🔄 Convertendo formato legado (analysisResults) para novo (subparcelas)")
                subparcelas_dict = {}
                for idx, result in enumerate(imported_data['analysisResults'], 1):
                    subparcela_id = f"sub_{idx}"
                    subparcelas_dict[subparcela_id] = {
                        'nome': result.get('subparcela', f'Sub {idx}'),
                        'image_path': result.get('image_path', ''),
                        'especies': result.get('especies', []),
                        'cobertura_total': result.get('cobertura_total', 0),
                        'area_descoberta': result.get('area_descoberta', 0)
                    }
                imported_data['subparcelas'] = subparcelas_dict
                
                # Converter especies para especies_unificadas se necessário
                if 'especies' in imported_data and 'especies_unificadas' not in imported_data:
                    imported_data['especies_unificadas'] = imported_data['especies']
            
            parcela_name = imported_data['parcela']
            print(f"✓ Parcela: {parcela_name}")
            print(f"✓ Subparcelas no JSON: {len(imported_data.get('subparcelas', {}))}")
            
            # Criar diretório de uploads para esta parcela
            upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], parcela_name)
            os.makedirs(upload_dir, exist_ok=True)
            
            # Copiar imagens para local permanente
            # Procurar QUALQUER diretório com imagens (subparcelas, images, especies, etc)
            image_mapping = {}  # filename -> new_path
            images_found = False
            
            # Buscar recursivamente por todos os arquivos de imagem
            print(f"🔍 Procurando imagens em {temp_dir}...")
            for root, dirs, files in os.walk(temp_dir):
                for filename in files:
                    # Verificar se é arquivo de imagem
                    if filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp')):
                        src = os.path.join(root, filename)
                        dst = os.path.join(upload_dir, filename)
                        shutil.copy2(src, dst)
                        image_mapping[filename] = dst
                        print(f"   ✓ {filename}")
                        images_found = True
            
            if images_found:
                print(f"✓ {len(image_mapping)} imagens copiadas")
            else:
                print(f"⚠️ Nenhuma imagem encontrada no ZIP")
            
            # Atualizar paths das imagens nas subparcelas
            for subparcela_id, subparcela in imported_data['subparcelas'].items():
                old_path = subparcela.get('image_path', '')
                filename = os.path.basename(old_path)
                if filename in image_mapping:
                    subparcela['image_path'] = image_mapping[filename]
            
            # Reconstruir lista de images com estrutura correta
            images_list = []
            for idx, (filename, new_path) in enumerate(sorted(image_mapping.items()), 1):
                images_list.append({
                    'filename': filename,
                    'path': new_path,
                    'subparcela': idx
                })
            
            # Restaurar dados na memória
            if parcela_name not in analysis_data['parcelas']:
                analysis_data['parcelas'][parcela_name] = {
                    'images': [],
                    'subparcelas': {}
                }
            
            analysis_data['parcelas'][parcela_name]['subparcelas'] = imported_data['subparcelas']
            analysis_data['parcelas'][parcela_name]['images'] = images_list
            
            # Restaurar espécies unificadas (já vem no formato correto, sem nível extra de aninhamento)
            analysis_data['especies_unificadas'][parcela_name] = imported_data.get('especies_unificadas', {})
            
            print(f"✓ Análise importada: {parcela_name}")
            print(f"✓ Subparcelas: {len(imported_data['subparcelas'])}")
            print(f"✓ Espécies: {len(imported_data.get('especies_unificadas', {}))}")
            
            # Limpar diretório temporário
            shutil.rmtree(temp_dir, ignore_errors=True)
            
            # Preparar dados de resposta completos para restaurar a interface
            analysis_results = []
            for subparcela_id, subparcela_data in sorted(imported_data['subparcelas'].items()):
                analysis_results.append({
                    'subparcela': subparcela_data.get('nome', f'Sub {len(analysis_results) + 1}'),
                    'image_path': subparcela_data.get('image_path', ''),
                    'especies': subparcela_data.get('especies', []),
                    'cobertura_total': subparcela_data.get('cobertura_total', 0),
                    'area_descoberta': subparcela_data.get('area_descoberta', 0)
                })
            
            subparcelas_list = [
                {'name': f"Subparcela {i+1}", 'path': r['image_path']}
                for i, r in enumerate(analysis_results)
            ]
            
            return jsonify({
                'success': True,
                'message': f'Análise "{parcela_name}" importada com sucesso',
                'parcela': parcela_name,
                'metadata': imported_data.get('metadata', {}),
                'analysis_results': analysis_results,
                'especies': imported_data.get('especies_unificadas', {}),
                'subparcelas': subparcelas_list
            })
            
        except Exception as e:
            # Limpar em caso de erro
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir, ignore_errors=True)
            raise e
            
    except Exception as e:
        print(f"Erro ao importar análise completa: {e}")
        return jsonify({'error': str(e)}), 500

# ====== ENDPOINTS DE FOTOS DE ESPÉCIES ======

# Estrutura para armazenar fotos das espécies
species_photos = {}  # {apelido_original: [{id, filename, url, uploaded_at}]}

@app.route('/api/especies/<apelido>/photos', methods=['GET'])
def get_species_photos(apelido):
    """Retorna lista de fotos de uma espécie"""
    photos = species_photos.get(apelido, [])
    return jsonify({
        'success': True,
        'photos': photos
    })

@app.route('/api/especies/<apelido>/photos', methods=['POST'])
def upload_species_photos(apelido):
    """Upload de fotos de uma espécie"""
    try:
        if 'photos' not in request.files:
            return jsonify({'error': 'Nenhuma foto enviada'}), 400

        files = request.files.getlist('photos')

        if not files:
            return jsonify({'error': 'Nenhuma foto selecionada'}), 400

        # Criar diretório para fotos de espécies se não existir
        photos_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'species_photos', secure_filename(apelido))
        os.makedirs(photos_dir, exist_ok=True)

        uploaded_photos = []

        for file in files:
            if file and file.filename:
                # Gerar nome único para o arquivo
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                original_filename = secure_filename(file.filename)
                filename = f"{timestamp}_{original_filename}"
                filepath = os.path.join(photos_dir, filename)

                # Salvar arquivo
                file.save(filepath)

                # Criar URL relativa
                relative_path = filepath.replace('static/', '').replace('\\', '/')
                photo_url = f'/static/{relative_path}'

                # Criar registro da foto
                photo_record = {
                    'id': f"{apelido}_{timestamp}_{original_filename}",
                    'filename': original_filename,
                    'url': photo_url,
                    'uploaded_at': datetime.now().isoformat()
                }

                uploaded_photos.append(photo_record)

        # Adicionar fotos ao registro da espécie
        if apelido not in species_photos:
            species_photos[apelido] = []

        species_photos[apelido].extend(uploaded_photos)

        # Salvar em arquivo JSON para persistência
        photos_data_file = os.path.join(app.config['UPLOAD_FOLDER'], 'species_photos_data.json')
        with open(photos_data_file, 'w', encoding='utf-8') as f:
            json.dump(species_photos, f, ensure_ascii=False, indent=2)

        print(f"✓ {len(uploaded_photos)} foto(s) adicionada(s) para {apelido}")

        return jsonify({
            'success': True,
            'uploaded': len(uploaded_photos),
            'photos': species_photos[apelido]
        })

    except Exception as e:
        print(f"Erro ao fazer upload de fotos: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/especies/<apelido>/photos/<photo_id>', methods=['DELETE'])
def delete_species_photo(apelido, photo_id):
    """Remove uma foto de uma espécie"""
    try:
        if apelido not in species_photos:
            return jsonify({'error': 'Espécie não encontrada'}), 404

        # Encontrar e remover a foto
        photo_to_remove = None
        for photo in species_photos[apelido]:
            if photo['id'] == photo_id:
                photo_to_remove = photo
                break

        if not photo_to_remove:
            return jsonify({'error': 'Foto não encontrada'}), 404

        # Remover arquivo físico
        file_path = photo_to_remove['url'].replace('/static/', 'static/')
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"✓ Arquivo removido: {file_path}")

        # Remover do registro
        species_photos[apelido] = [p for p in species_photos[apelido] if p['id'] != photo_id]

        # Se não houver mais fotos, limpar diretório
        if not species_photos[apelido]:
            photos_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'species_photos', secure_filename(apelido))
            if os.path.exists(photos_dir):
                shutil.rmtree(photos_dir)
                print(f"✓ Diretório removido: {photos_dir}")
            del species_photos[apelido]

        # Salvar alterações
        photos_data_file = os.path.join(app.config['UPLOAD_FOLDER'], 'species_photos_data.json')
        with open(photos_data_file, 'w', encoding='utf-8') as f:
            json.dump(species_photos, f, ensure_ascii=False, indent=2)

        return jsonify({
            'success': True,
            'message': 'Foto removida com sucesso'
        })

    except Exception as e:
        print(f"Erro ao remover foto: {e}")
        return jsonify({'error': str(e)}), 500

# Carregar fotos salvas na inicialização
def load_species_photos():
    """Carrega fotos das espécies do arquivo JSON"""
    global species_photos
    photos_data_file = os.path.join(app.config['UPLOAD_FOLDER'], 'species_photos_data.json')

    if os.path.exists(photos_data_file):
        try:
            with open(photos_data_file, 'r', encoding='utf-8') as f:
                species_photos = json.load(f)
            print(f"✓ Fotos de espécies carregadas: {len(species_photos)} espécie(s)")
        except Exception as e:
            print(f"Erro ao carregar fotos: {e}")
            species_photos = {}
    else:
        species_photos = {}

# Carregar fotos na inicialização
load_species_photos()

@app.route('/export_pdf', methods=['POST'])
def export_pdf():
    """Exporta análise completa para PDF"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from io import BytesIO
        
        data = request.json
        parcela_nome = data.get('parcela', 'Parcela')
        especies = data.get('especies', {})
        analytics = data.get('analytics', {})
        
        # Criar PDF em memória
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        
        story = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2196F3'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        story.append(Paragraph(f"Relatório de Análise de Vegetação Herbácea", title_style))
        story.append(Paragraph(f"<b>Parcela:</b> {parcela_nome}", styles['Heading2']))
        story.append(Spacer(1, 20))
        
        # Análises Ecológicas
        story.append(Paragraph("Análises Ecológicas", styles['Heading2']))
        
        analytics_data = [
            ['Índice', 'Valor', 'Interpretação'],
            ['Diversidade de Shannon (H\')', f"{analytics.get('diversity', 0):.3f}", 'Indica diversidade de espécies'],
            ['Riqueza de Espécies (S)', str(analytics.get('richness', 0)), 'Número total de espécies'],
            ['Equitabilidade de Pielou (J\')', f"{analytics.get('eveness', 0):.3f}", 'Uniformidade da distribuição'],
            ['Dominância de Simpson (D)', f"{analytics.get('simpson', 0):.3f}", 'Probabilidade de dominância']
        ]
        
        t = Table(analytics_data, colWidths=[6*cm, 4*cm, 8*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(t)
        story.append(Spacer(1, 30))
        
        # Tabela de Espécies
        story.append(Paragraph("Lista de Espécies Identificadas", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        especies_data = [['Apelido', 'Gênero', 'Espécie', 'Família', 'Cobertura (%)', 'Ocorrências']]
        
        for esp_nome, esp_info in sorted(especies.items(), key=lambda x: x[1].get('cobertura', 0), reverse=True):
            especies_data.append([
                esp_info.get('apelido_usuario', esp_nome),
                esp_info.get('genero', '-'),
                esp_info.get('especie', '-'),
                esp_info.get('familia', '-'),
                f"{esp_info.get('cobertura', 0):.2f}",
                str(esp_info.get('ocorrencias', 0))
            ])
        
        t_especies = Table(especies_data, colWidths=[4*cm, 3*cm, 3*cm, 3*cm, 2.5*cm, 2.5*cm])
        t_especies.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9)
        ]))
        
        story.append(t_especies)
        
        # Construir PDF
        doc.build(story)
        
        # Retornar PDF
        buffer.seek(0)
        return send_file(buffer, 
                        mimetype='application/pdf',
                        as_attachment=True,
                        download_name=f'{parcela_nome}_relatorio_completo.pdf')
    
    except Exception as e:
        print(f"Erro ao gerar PDF: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/export_zip', methods=['POST'])
def export_zip():
    """Exporta pacote completo com Excel, PDF, fotos e JSON"""
    try:
        import zipfile
        from io import BytesIO
        
        data = request.json
        parcela_nome = data.get('parcela', 'Parcela')
        especies = data.get('especies', {})
        analysis_results = data.get('analysisResults', [])
        analytics = data.get('analytics', {})
        
        # Criar ZIP em memória
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # 1. Adicionar JSON com todos os dados
            # Converter analysisResults para formato subparcelas
            subparcelas_dict = {}
            for idx, result in enumerate(analysis_results, 1):
                subparcela_id = f"sub_{idx}"
                subparcelas_dict[subparcela_id] = {
                    'nome': result.get('subparcela', f'Sub {idx}'),
                    'image_path': result.get('image_path', ''),
                    'especies': result.get('especies', []),
                    'cobertura_total': result.get('cobertura_total', 0),
                    'area_descoberta': result.get('area_descoberta', 0)
                }
            
            json_data = {
                'parcela': parcela_nome,
                'especies_unificadas': especies,
                'subparcelas': subparcelas_dict,
                'analytics': analytics,
                'data_exportacao': datetime.now().isoformat(),
                'metadata': {
                    'versao': '2.0',
                    'total_subparcelas': len(analysis_results)
                }
            }
            zip_file.writestr(f'{parcela_nome}_dados.json', 
                            json.dumps(json_data, indent=2, ensure_ascii=False))
            
            # 2. Adicionar Excel (chamando a função existente)
            # TODO: Implementar geração de Excel com analytics
            
            # 3. Adicionar fotos das subparcelas (APENAS as da análise atual)
            upload_folder = app.config['UPLOAD_FOLDER']
            parcela_folder = os.path.join(upload_folder, parcela_nome)
            
            # Criar conjunto com apenas os nomes de arquivo das subparcelas na análise
            current_images = set()
            for result in analysis_results:
                image_path = result.get('image_path', '')
                if image_path:
                    filename = os.path.basename(image_path)
                    current_images.add(filename)
            
            print(f"📁 Exportando {len(current_images)} imagens da análise atual")
            
            if os.path.exists(parcela_folder) and current_images:
                for filename in os.listdir(parcela_folder):
                    # Incluir APENAS as imagens que estão na análise atual
                    if filename in current_images and filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                        file_path = os.path.join(parcela_folder, filename)
                        zip_file.write(file_path, f'images/{filename}')
                        print(f"   ✓ {filename}")
            
            # 4. Adicionar fotos das espécies
            species_photos_folder = os.path.join(upload_folder, 'species_photos')
            
            if os.path.exists(species_photos_folder):
                for species_folder in os.listdir(species_photos_folder):
                    species_path = os.path.join(species_photos_folder, species_folder)
                    if os.path.isdir(species_path):
                        for filename in os.listdir(species_path):
                            if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                                file_path = os.path.join(species_path, filename)
                                zip_file.write(file_path, f'especies/{species_folder}/{filename}')
            
            # 5. Adicionar README
            readme_content = f"""
# Pacote de Análise de Vegetação Herbácea
## {parcela_nome}

Data de Exportação: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}

### Conteúdo do Pacote:
- **{parcela_nome}_dados.json**: Todos os dados da análise em formato JSON
- **subparcelas/**: Fotos das subparcelas analisadas
- **especies/**: Fotos das espécies identificadas (organizadas por espécie)

### Análises Ecológicas:
- Diversidade de Shannon: {analytics.get('diversity', 0):.3f}
- Riqueza de Espécies: {analytics.get('richness', 0)}
- Equitabilidade: {analytics.get('eveness', 0):.3f}
- Dominância de Simpson: {analytics.get('simpson', 0):.3f}

### Espécies Identificadas: {len(especies)}

Para reimportar esta análise no sistema, use a opção "Importar ZIP" e selecione este arquivo.
"""
            zip_file.writestr('README.txt', readme_content)
        
        # Retornar ZIP
        zip_buffer.seek(0)
        return send_file(zip_buffer,
                        mimetype='application/zip',
                        as_attachment=True,
                        download_name=f'{parcela_nome}_pacote_completo.zip')
    
    except Exception as e:
        print(f"Erro ao gerar ZIP: {e}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
