import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Análise REFINADA das 10 subparcelas da Parcela 9
# Identificando TODAS as formas vegetais, incluindo leguminosas, arbustos, e diferentes herbáceas

parcela9_refined = [
    # Subparcela 1 (TimePhoto_20251107_133034.jpg - índice 4107)
    # Vegetação densa com gramíneas verdes e material seco
    (9, 1, 1, 'Gramínea Verde Larga', 25, 38, 'Erva'),
    (9, 1, 2, 'Gramínea Seca Palha', 40, 28, 'Erva'),
    (9, 1, 3, 'Gramínea Fina Verde', 20, 22, 'Erva'),
    (9, 1, 4, 'Herbácea Folha Larga', 10, 15, 'Erva'),
    (9, 1, 5, 'Solo/Serapilheira', 5, 0, '-'),

    # Subparcela 2 (TimePhoto_20251107_133249.jpg - índice 4108)
    # Dominada por material seco com gramíneas verdes esparsas
    (9, 2, 1, 'Gramínea Seca Palha', 55, 32, 'Erva'),
    (9, 2, 2, 'Gramínea Verde Larga', 20, 42, 'Erva'),
    (9, 2, 3, 'Gramínea Fina Verde', 15, 20, 'Erva'),
    (9, 2, 4, 'Herbácea Folha Larga', 5, 12, 'Erva'),
    (9, 2, 5, 'Solo/Serapilheira', 5, 0, '-'),

    # Subparcela 3 (TimePhoto_20251107_133446.jpg - índice 4140)
    # IMPORTANTE: Presença clara de LEGUMINOSA com folhas compostas
    (9, 3, 1, 'Gramínea Seca Palha', 45, 30, 'Erva'),
    (9, 3, 2, 'Gramínea Verde Larga', 20, 35, 'Erva'),
    (9, 3, 3, 'Leguminosa Folha Composta', 15, 25, 'Erva'),
    (9, 3, 4, 'Gramínea Fina Verde', 10, 18, 'Erva'),
    (9, 3, 5, 'Herbácea Folha Larga', 5, 10, 'Erva'),
    (9, 3, 6, 'Solo/Serapilheira', 5, 0, '-'),

    # Subparcela 4 (TimePhoto_20251107_133807.jpg - índice 4119)
    # Solo mais exposto, vegetação esparsa com gramíneas baixas
    (9, 4, 1, 'Gramínea Seca Palha', 35, 25, 'Erva'),
    (9, 4, 2, 'Solo/Serapilheira', 30, 0, '-'),
    (9, 4, 3, 'Gramínea Baixa Roseta', 15, 8, 'Erva'),
    (9, 4, 4, 'Gramínea Fina Verde', 10, 14, 'Erva'),
    (9, 4, 5, 'Gramínea Verde Larga', 5, 20, 'Erva'),
    (9, 4, 6, 'Herbácea Folha Larga', 5, 10, 'Erva'),

    # Subparcela 5 (TimePhoto_20251107_134123.jpg - índice 4121)
    # Dominância forte de material seco
    (9, 5, 1, 'Gramínea Seca Palha', 65, 28, 'Erva'),
    (9, 5, 2, 'Gramínea Verde Larga', 20, 32, 'Erva'),
    (9, 5, 3, 'Gramínea Fina Verde', 10, 16, 'Erva'),
    (9, 5, 4, 'Herbácea Folha Larga', 3, 8, 'Erva'),
    (9, 5, 5, 'Solo/Serapilheira', 2, 0, '-'),

    # Subparcela 6 (TimePhoto_20251107_134343.jpg - índice 4122)
    # Mistura equilibrada de gramíneas verdes e secas
    (9, 6, 1, 'Gramínea Seca Palha', 40, 32, 'Erva'),
    (9, 6, 2, 'Gramínea Verde Larga', 30, 40, 'Erva'),
    (9, 6, 3, 'Gramínea Fina Verde', 18, 22, 'Erva'),
    (9, 6, 4, 'Herbácea Folha Larga', 8, 15, 'Erva'),
    (9, 6, 5, 'Solo/Serapilheira', 4, 0, '-'),

    # Subparcela 7 (TimePhoto_20251107_134654.jpg - índice 4127)
    # Gramíneas secas dominantes com algumas verdes
    (9, 7, 1, 'Gramínea Seca Palha', 45, 30, 'Erva'),
    (9, 7, 2, 'Gramínea Verde Larga', 25, 36, 'Erva'),
    (9, 7, 3, 'Gramínea Fina Verde', 15, 20, 'Erva'),
    (9, 7, 4, 'Herbácea Folha Larga', 5, 12, 'Erva'),
    (9, 7, 5, 'Solo/Serapilheira', 10, 0, '-'),

    # Subparcela 8 (TimePhoto_20251107_135003.jpg - índice 4126)
    # Boa presença de gramíneas verdes vigorosas
    (9, 8, 1, 'Gramínea Verde Larga', 40, 44, 'Erva'),
    (9, 8, 2, 'Gramínea Seca Palha', 30, 28, 'Erva'),
    (9, 8, 3, 'Gramínea Fina Verde', 15, 20, 'Erva'),
    (9, 8, 4, 'Herbácea Folha Larga', 10, 14, 'Erva'),
    (9, 8, 5, 'Leguminosa Folha Composta', 3, 18, 'Erva'),
    (9, 8, 6, 'Solo/Serapilheira', 2, 0, '-'),

    # Subparcela 9 (TimePhoto_20251107_135459.jpg - índice 4155)
    # Gramíneas verdes altas e vigorosas
    (9, 9, 1, 'Gramínea Verde Larga', 45, 48, 'Erva'),
    (9, 9, 2, 'Gramínea Seca Palha', 30, 32, 'Erva'),
    (9, 9, 3, 'Gramínea Fina Verde', 15, 22, 'Erva'),
    (9, 9, 4, 'Herbácea Folha Larga', 5, 12, 'Erva'),
    (9, 9, 5, 'Leguminosa Folha Composta', 3, 20, 'Erva'),
    (9, 9, 6, 'Solo/Serapilheira', 2, 0, '-'),

    # Subparcela 10 (TimePhoto_20251107_135623.jpg - índice 4136)
    # Gramíneas verdes altas dominantes
    (9, 10, 1, 'Gramínea Verde Larga', 50, 50, 'Erva'),
    (9, 10, 2, 'Gramínea Seca Palha', 25, 30, 'Erva'),
    (9, 10, 3, 'Gramínea Fina Verde', 15, 24, 'Erva'),
    (9, 10, 4, 'Herbácea Folha Larga', 6, 14, 'Erva'),
    (9, 10, 5, 'Leguminosa Folha Composta', 2, 22, 'Erva'),
    (9, 10, 6, 'Solo/Serapilheira', 2, 0, '-'),
]

# Criar nova planilha
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Parcela 9 - Herbáceas"

# Cabeçalho
headers = ['Parcela', 'Subparcela', 'Índice', 'Nome da Espécie', 'Cobertura (%)', 'Altura (cm)', 'Forma de Vida']
ws.append(headers)

# Formatar cabeçalho
header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Adicionar dados
for data in parcela9_refined:
    parcela, sub, ind, nome, cobertura, altura, forma = data
    ws.append([
        parcela if ind == 1 else None,
        sub if ind == 1 else None,
        ind,
        nome,
        cobertura,
        altura,
        forma
    ])

# Ajustar largura das colunas
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 28
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 14

# Criar planilha de resumo
ws_resumo = wb.create_sheet("Resumo por Espécie")
ws_resumo.append(['Espécie', 'Nº Ocorrências', 'Cobertura Média (%)', 'Altura Média (cm)', 'Forma de Vida'])

# Calcular resumo por espécie
especies_dict = {}
for data in parcela9_refined:
    _, _, _, nome, cobertura, altura, forma = data
    if nome not in especies_dict:
        especies_dict[nome] = {'ocorrencias': 0, 'cobertura_total': 0, 'altura_total': 0, 'forma': forma}
    especies_dict[nome]['ocorrencias'] += 1
    especies_dict[nome]['cobertura_total'] += cobertura
    especies_dict[nome]['altura_total'] += altura

for especie, dados in sorted(especies_dict.items()):
    n_ocorrencias = dados['ocorrencias']
    cobertura_media = round(dados['cobertura_total'] / n_ocorrencias, 1)
    altura_media = round(dados['altura_total'] / n_ocorrencias, 1) if dados['altura_total'] > 0 else 0
    ws_resumo.append([especie, n_ocorrencias, cobertura_media, altura_media, dados['forma']])

# Formatar cabeçalho do resumo
for cell in ws_resumo[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Ajustar largura das colunas do resumo
ws_resumo.column_dimensions['A'].width = 28
ws_resumo.column_dimensions['B'].width = 16
ws_resumo.column_dimensions['C'].width = 18
ws_resumo.column_dimensions['D'].width = 16
ws_resumo.column_dimensions['E'].width = 14

# Salvar arquivo
output_file = r'C:\Users\diogo\Documents\TRABALHO\dossel\Parcela9_Herbaceas_REFINADO.xlsx'
wb.save(output_file)

print(f'Analise refinada concluida!')
print(f'Arquivo salvo: {output_file}')
print(f'\nEstatísticas:')
print(f'  - Total de registros: {len(parcela9_refined)}')
print(f'  - Total de espécies identificadas: {len(especies_dict)}')
print(f'  - Subparcelas analisadas: 10')
print(f'\nEspécies encontradas:')
for i, especie in enumerate(sorted(especies_dict.keys()), 1):
    print(f'  {i}. {especie} ({especies_dict[especie]["ocorrencias"]} ocorrências)')
